"""实验附件：本地 SQLite 内容库 + 实验引用。"""

from __future__ import annotations

import csv
import math
import zipfile
from io import BytesIO
from pathlib import Path

from flask import Blueprint, g, jsonify, request, send_file


api_attachment_bp = Blueprint("api_attachment", __name__)

_ALLOWED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".pdf", ".csv", ".txt", ".xlsx", ".mp3", ".ogg", ".wav", ".m4a", ".aac", ".flac"}
_MAX_ATTACHMENT_BYTES = 16 * 1024 * 1024
_PREVIEW_MAX_ROWS = 200
_PREVIEW_MAX_COLS = 50
_PREVIEW_MAX_TEXT_CHARS = 200_000


def _notify_attachment_change(exp_id: str) -> None:
    from routes.api_agent import publish_resource_change
    publish_resource_change(exp_id, "edited", request.headers.get("X-Exdiary-Client-Id", ""))


def _attachment_links(experiment: dict) -> list[dict]:
    links = experiment.get("attachments") or []
    return [link for link in links if isinstance(link, dict) and link.get("sha256")]


def _display_attachment(link: dict) -> dict:
    meta = g.attachment_store.meta(link["sha256"]) or {}
    return {
        **link,
        "size": int(link.get("size") or meta.get("size") or 0),
        "mime": link.get("mime") or meta.get("mime") or "",
        "sync_state": meta.get("sync_state", "pending"),
        "has_content": bool(meta.get("has_content")),
        "url": f"/api/attachments/{link['sha256']}",
    }


def store_uploaded_attachment(file) -> dict:
    """校验并写入附件本体；聊天和实验详情共用同一套限制。"""
    if not file or not file.filename:
        raise ValueError("请选择文件")
    name = Path(file.filename).name
    suffix = Path(name).suffix.lower()
    mime = (file.mimetype or "").lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise ValueError("仅支持图片、文档、表格和常见音频文件")
    content = file.read(_MAX_ATTACHMENT_BYTES + 1)
    if not content:
        raise ValueError("文件为空")
    if len(content) > _MAX_ATTACHMENT_BYTES:
        raise OverflowError("单个附件不能超过 16 MB")
    return g.attachment_store.put(content, name=name, mime=mime)


@api_attachment_bp.route("/experiments/<exp_id>/attachments")
def api_experiment_attachments(exp_id: str):
    experiment = g.exp_repo.load(exp_id)
    if not experiment:
        return jsonify({"ok": False, "error": "实验不存在"}), 404
    return jsonify({"ok": True, "attachments": [_display_attachment(link) for link in _attachment_links(experiment)]})


@api_attachment_bp.route("/attachments", methods=["POST"])
def api_upload_attachment():
    exp_id = (request.form.get("experiment_id") or "").strip()
    file = request.files.get("file")
    if not exp_id or not file or not file.filename:
        return jsonify({"ok": False, "error": "请选择文件并指定实验"}), 400
    experiment = g.exp_repo.load(exp_id)
    if not experiment:
        return jsonify({"ok": False, "error": "实验不存在"}), 404

    name = Path(file.filename).name
    try:
        meta = store_uploaded_attachment(file)
    except OverflowError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 413
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    mime = meta.get("mime", "")
    links = _attachment_links(experiment)
    title = (request.form.get("title") or request.form.get("caption") or name).strip()[:200]
    description = (request.form.get("description") or "").strip()[:1000]
    kind = (request.form.get("kind") or "attachment").strip()[:40]
    link = {
        "sha256": meta["sha256"], "name": name, "mime": mime,
        "size": meta["size"], "title": title, "description": description, "kind": kind,
        "created_at": meta.get("created_at", ""),
    }
    existing = next((item for item in links if item["sha256"] == meta["sha256"]), None)
    if existing:
        existing.update({key: value for key, value in link.items() if value})
    else:
        links.append(link)
    experiment["attachments"] = links
    expected = request.form.get("expected_revision")
    if expected is not None and hasattr(g.exp_repo, "save_if_revision"):
        try:
            saved = g.experiment_svc.save_and_update_refs_if_revision(
                exp_id, experiment, int(expected), source="user"
            )
        except ValueError:
            return jsonify({"ok": False, "error": "版本号无效"}), 400
        if not saved.get("ok"):
            return jsonify({"ok": False, "error": "内容已在另一窗口更新，请刷新后再操作。",
                            "revision": saved.get("revision")}), 409
        _notify_attachment_change(exp_id)
        return jsonify({"ok": True, "attachment": _display_attachment(link),
                        "revision": saved["revision"]})
    g.experiment_svc.save_with_log(exp_id, experiment, "user")
    _notify_attachment_change(exp_id)
    return jsonify({"ok": True, "attachment": _display_attachment(link),
                    "revision": (g.exp_repo.load(exp_id) or {}).get("revision", 0)})


@api_attachment_bp.route("/experiments/<exp_id>/attachments/<sha256>", methods=["DELETE"])
def api_remove_attachment_link(exp_id: str, sha256: str):
    experiment = g.exp_repo.load(exp_id)
    if not experiment:
        return jsonify({"ok": False, "error": "实验不存在"}), 404
    links = _attachment_links(experiment)
    kept = [link for link in links if link["sha256"] != sha256]
    if len(kept) == len(links):
        return jsonify({"ok": False, "error": "附件引用不存在"}), 404
    experiment["attachments"] = kept
    expected = request.args.get("expected_revision")
    if expected is not None and hasattr(g.exp_repo, "save_if_revision"):
        try:
            saved = g.experiment_svc.save_and_update_refs_if_revision(
                exp_id, experiment, int(expected), source="user"
            )
        except ValueError:
            return jsonify({"ok": False, "error": "版本号无效"}), 400
        if not saved.get("ok"):
            return jsonify({"ok": False, "error": "内容已在另一窗口更新，请刷新后再操作。",
                            "revision": saved.get("revision")}), 409
        _notify_attachment_change(exp_id)
        return jsonify({"ok": True, "revision": saved["revision"]})
    g.experiment_svc.save_with_log(exp_id, experiment, "user")
    _notify_attachment_change(exp_id)
    return jsonify({"ok": True, "revision": (g.exp_repo.load(exp_id) or {}).get("revision", 0)})


@api_attachment_bp.route("/attachments/<sha256>")
def api_download_attachment(sha256: str):
    content = g.attachment_store.get(sha256)
    meta = g.attachment_store.meta(sha256)
    if content is None or meta is None:
        return jsonify({"ok": False, "error": "附件尚未下载到此设备"}), 404
    as_download = request.args.get("download") == "1"
    return send_file(BytesIO(content), mimetype=meta.get("mime") or "application/octet-stream",
                     as_attachment=as_download, download_name=meta.get("name") or sha256)


@api_attachment_bp.route("/attachments/<sha256>/preview")
def api_attachment_preview(sha256: str):
    """为应用内只读查看提供安全、受限的文本与表格预览。"""
    content = g.attachment_store.get(sha256)
    meta = g.attachment_store.meta(sha256)
    if content is None or meta is None:
        return jsonify({"ok": False, "error": "附件尚未下载到本机"}), 404
    name = meta.get("name") or sha256
    mime = (meta.get("mime") or "").lower()
    suffix = Path(name).suffix.lower()
    try:
        if suffix == ".xlsx" or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter, range_boundaries
            # 预览面向阅读：公式单元格展示 Excel/WPS 上次保存的计算结果，
            # 不把 =SUM(...) 等公式文本当作数据表内容显示。
            # 列宽、行高保存在工作表维度中，openpyxl 的只读模式不会提供它们。
            # 该操作只在用户主动打开预览时发生，表格窗口仍限制在 200×50 单元格。
            book = load_workbook(BytesIO(content), read_only=False, data_only=True, keep_links=False)
            sheet_name = request.args.get("sheet") or book.sheetnames[0]
            if sheet_name not in book.sheetnames:
                return jsonify({"ok": False, "error": "工作表不存在"}), 404
            sheet = book[sheet_name]
            try:
                start_row = max(1, int(request.args.get("start_row", 1)))
            except ValueError:
                start_row = 1
            rows = []
            cell_styles = {}
            max_col = min(sheet.max_column or 1, _PREVIEW_MAX_COLS)
            row_limit = min(sheet.max_row, start_row + _PREVIEW_MAX_ROWS - 1)
            for row in sheet.iter_rows(
                min_row=start_row, max_row=row_limit,
                max_col=max_col,
            ):
                values = []
                for cell in row:
                    values.append(_preview_cell(cell.value))
                    style = _preview_cell_style(cell)
                    if style:
                        cell_styles[cell.coordinate] = style
                rows.append(values)
            column_letters = [get_column_letter(index) for index in range(1, max_col + 1)]
            return jsonify({
                "ok": True, "kind": "xlsx", "name": name, "sheet": sheet_name,
                "sheets": [{"name": item.title, "rows": item.max_row, "columns": item.max_column}
                           for item in book.worksheets],
                "start_row": start_row, "columns": [get_column_letter(index)
                    for index in range(1, max_col + 1)],
                "rows": rows, "cell_styles": cell_styles,
                "column_widths": _preview_column_widths(sheet, column_letters),
                "row_heights": _preview_row_heights(sheet, start_row, row_limit),
                "merged_cells": _preview_merged_cells(
                    sheet, start_row, row_limit, max_col, range_boundaries
                ),
                "has_more": sheet.max_row > row_limit,
            })
        if suffix in (".csv", ".tsv") or mime in ("text/csv", "text/tab-separated-values"):
            text = content.decode("utf-8-sig", errors="replace")
            delimiter = "\t" if suffix == ".tsv" or mime == "text/tab-separated-values" else ","
            rows = []
            for row in csv.reader(text.splitlines(), delimiter=delimiter):
                rows.append(row[:_PREVIEW_MAX_COLS])
                if len(rows) >= _PREVIEW_MAX_ROWS:
                    break
            return jsonify({"ok": True, "kind": "table", "name": name, "rows": rows,
                            "has_more": len(text.splitlines()) > len(rows)})
        if suffix == ".txt" or mime.startswith("text/"):
            text = content.decode("utf-8-sig", errors="replace")
            return jsonify({"ok": True, "kind": "text", "name": name,
                            "content": text[:_PREVIEW_MAX_TEXT_CHARS],
                            "truncated": len(text) > _PREVIEW_MAX_TEXT_CHARS})
    except (ImportError, OSError, ValueError, zipfile.BadZipFile) as exc:
        return jsonify({"ok": False, "error": f"附件预览失败：{str(exc)[:160]}"}), 422
    return jsonify({"ok": False, "error": "该文件类型不提供结构化预览"}), 415


def _preview_cell(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if not math.isfinite(number):
            return str(value)
        if number.is_integer():
            return str(int(number))
        magnitude = abs(number)
        if magnitude >= 1_000_000_000 or 0 < magnitude < 0.0001:
            return f"{number:.4g}"
        return f"{number:.4f}".rstrip("0").rstrip(".")
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _preview_column_widths(sheet, letters: list[str]) -> dict[str, float]:
    default = sheet.sheet_format.defaultColWidth or 8.43
    widths = {}
    for letter in letters:
        dimension = sheet.column_dimensions.get(letter)
        width = dimension.width if dimension and dimension.width is not None else default
        # Excel 列宽的单位不是像素；这里按常用字体的近似换算，保留合理上下限。
        widths[letter] = round(max(28, min(520, float(width) * 7 + 5)), 1)
    return widths


def _preview_row_heights(sheet, start_row: int, end_row: int) -> dict[str, float]:
    default = sheet.sheet_format.defaultRowHeight or 15
    heights = {}
    for index in range(start_row, end_row + 1):
        dimension = sheet.row_dimensions.get(index)
        height = dimension.height if dimension and dimension.height is not None else default
        heights[str(index)] = round(max(16, min(360, float(height) * 96 / 72)), 1)
    return heights


def _preview_merged_cells(sheet, start_row: int, end_row: int, max_col: int, range_boundaries) -> list[dict]:
    """仅返回当前页完整可见的合并区域，前端据此生成 rowspan/colspan。"""
    merged = []
    for item in sheet.merged_cells.ranges:
        min_col, min_row, max_merge_col, max_merge_row = range_boundaries(str(item))
        if min_row < start_row or max_merge_row > end_row or min_col < 1 or max_merge_col > max_col:
            continue
        merged.append({
            "start_row": min_row, "start_col": min_col,
            "end_row": max_merge_row, "end_col": max_merge_col,
        })
    return sorted(merged, key=lambda item: (item["start_row"], item["start_col"]))


def _preview_cell_style(cell) -> dict:
    """将安全、有限的 openpyxl 样式转成浏览器可用的描述。"""
    if not cell.has_style:
        return {}
    result = {}
    font = cell.font
    font_data = {}
    if font.name:
        font_data["family"] = font.name
    if font.sz:
        font_data["size"] = round(float(font.sz) * 96 / 72, 1)
    if font.b:
        font_data["bold"] = True
    if font.i:
        font_data["italic"] = True
    if font.u:
        font_data["underline"] = True
    if font.strike:
        font_data["strike"] = True
    color = _preview_color(font.color)
    if color:
        font_data["color"] = color
    if font_data:
        result["font"] = font_data

    fill = cell.fill
    if fill.fill_type == "solid":
        color = _preview_color(fill.fgColor)
        if color:
            result["fill"] = color

    alignment = cell.alignment
    alignment_data = {}
    horizontal = {"centerContinuous": "center", "distributed": "justify"}.get(
        alignment.horizontal, alignment.horizontal
    )
    if horizontal and horizontal != "general":
        alignment_data["horizontal"] = horizontal
    if alignment.vertical:
        alignment_data["vertical"] = {"center": "middle"}.get(alignment.vertical, alignment.vertical)
    if alignment.wrap_text:
        alignment_data["wrap"] = True
    if alignment.text_rotation:
        alignment_data["rotation"] = int(alignment.text_rotation)
    if alignment_data:
        result["alignment"] = alignment_data

    borders = {}
    for side_name in ("left", "right", "top", "bottom"):
        border = _preview_border(getattr(cell.border, side_name))
        if border:
            borders[side_name] = border
    if borders:
        result["borders"] = borders
    return result


def _preview_border(side) -> str:
    if not side or not side.style:
        return ""
    style = {"dashed": "dashed", "dashDot": "dashed", "dashDotDot": "dashed",
             "dotted": "dotted", "double": "double"}.get(side.style, "solid")
    width = {"medium": 2, "mediumDashed": 2, "mediumDashDot": 2,
             "mediumDashDotDot": 2, "thick": 3}.get(side.style, 1)
    return f"{width}px {style} {_preview_color(side.color) or '#222'}"


def _preview_color(color) -> str:
    if not color:
        return ""
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        return f"#{color.rgb[-6:]}"
    if color.type == "indexed" and color.indexed is not None:
        try:
            from openpyxl.styles.colors import COLOR_INDEX
            value = COLOR_INDEX[color.indexed]
            return f"#{value[-6:]}" if value else ""
        except (IndexError, TypeError):
            return ""
    return ""

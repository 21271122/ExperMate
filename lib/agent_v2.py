"""
Exdiary Agent v2 — 基于 Tool Calling 的对话式实验记录系统

LLM 自主决策流程，Python 仅执行工具和注入 Schema 状态。
"""

import json, re, traceback
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Generator
from zoneinfo import ZoneInfo

from openai import APITimeoutError

from lib.logger import get_logger
from lib.core.agent_tools import (
    TOOL_LOAD_REFERENCE, TOOL_SEARCH_EXPERIMENTS, TOOL_UPDATE_SCHEMA,
    TOOL_GENERATE_RECORD, TOOL_START_RECORD_THREAD,
    TOOL_END_THREAD, TOOL_START_ANALYZE_THREAD, TOOL_SELECT_EXPERIMENTS,
    TOOL_GENERATE_ANALYSIS, TOOL_READ_ANALYSIS, TOOL_READ_UPDATE_LOG,
    TOOL_MODIFY_EXPERIMENT, TOOL_MANAGE_ARCHIVE, TOOL_MANAGE_CATEGORY, TOOL_READ_EXPERIMENT,
    TOOL_LIST_EXPERIMENTS, TOOL_SEARCH_CHAT_HISTORY, TOOL_READ_CHAT_HISTORY,
    TOOL_LIST_CHAT_SESSIONS, TOOL_BROWSE_CHAT_HISTORY,
    TOOL_SEARCH_ATTACHMENTS, TOOL_READ_ATTACHMENT, TOOL_MANAGE_ATTACHMENT,
    TOOL_MANAGE_MUSIC,
    TOOLS_OPENAI_FORMAT,
)
from lib.core.prompts import build_system_prompt
from lib.core.schema import DEFAULT_CONTEXT
from lib.experiment_ids import is_experiment_id
from lib.repositories.sqlite_schema import FIELD_DEFAULTS


# 临时测试窗口：达到 1 万 token 后压缩，仅保留最近约 3000 token 原文。
_DEFAULT_CONTEXT_COMPRESSION_TRIGGER_TOKENS = 300_000
_DEFAULT_CONTEXT_COMPRESSION_CHUNK_TOKENS = 260_000
_CONTEXT_COMPRESSION_REQUEST_TIMEOUT_SECONDS = 120


class ChildContext:
    """子 Agent 标记。仅子 Agent 实例时有效。"""
    __slots__ = ('is_child', 'is_legacy', 'exp_id', 'initial_history_len', 'agent_role')
    def __init__(self):
        self.is_child = False
        self.is_legacy = False
        self.exp_id = None
        self.initial_history_len = 0
        self.agent_role = None


class ThreadState:
    """线程状态。"""
    __slots__ = ('id', 'type', 'pending_start', 'current_turn_user_idx', 'last_ended_id')
    def __init__(self):
        self.id = None
        self.type = None
        self.pending_start = None
        self.current_turn_user_idx = -1
        self.last_ended_id = None


# Tool definitions, SYSTEM_PROMPT, and DEFAULT_CONTEXT migrated to lib/core/

def merge_context(context: dict, fields: dict) -> dict:
    """增量合并。简单字段覆盖；数组追加去重；嵌套对象递归合并。"""
    for key, value in fields.items():
        if key not in context:
            continue
        existing = context[key]
        if isinstance(existing, list) and isinstance(value, list):
            if not value:
                context[key] = []
            else:
                for item in value:
                    if isinstance(item, str) and item in existing:
                        continue
                    existing.append(item)
        elif isinstance(existing, dict) and isinstance(value, dict):
            if not value:
                context[key] = {}
            else:
                for sk, sv in value.items():
                    if isinstance(existing.get(sk), list) and isinstance(sv, list):
                        for i in sv:
                            if i not in existing[sk]:
                                existing[sk].append(i)
                    elif sv not in (None, ""):
                        existing[sk] = sv
        else:
            context[key] = value
    return context


def _is_filled(val) -> bool:
    """检查单个字段是否有值"""
    if val is None:
        return False
    if isinstance(val, list):
        return len(val) > 0
    if isinstance(val, dict):
        return any(v for v in val.values() if v)
    if isinstance(val, str):
        return val.strip() != ""
    return bool(val)


def _brief(val) -> str:
    """字段值的简短描述"""
    if isinstance(val, list):
        return f"{len(val)}项" if val else "空"
    if isinstance(val, dict):
        has = sum(1 for v in val.values() if v)
        return f"{has}子字段" if has else "空"
    if isinstance(val, str):
        return val[:15] + ("..." if len(val) > 15 else "")
    return "有" if val else "空"


def _build_preview(loop: "AgentLoop") -> dict:
    """确定性构建：从 context 直接构造待保存的实验记录，不调 LLM。
    parse_notes 失败时的退化兜底——保证输出结构完整，但不做语义推断和补全。"""
    ctx = loop._schema_context or {}
    return {
        "id": loop.store.next_id(),
        "title": ctx.get("title", ""),
        "date": ctx.get("date", ""),
        "experimenter": ctx.get("experimenter", ""),
        "status": ctx.get("status", "planned"),
        "tags": ctx.get("tags", []),
        "purpose": ctx.get("purpose", ""),
        "materials": ctx.get("materials", []),
        "equipment": ctx.get("equipment", []),
        "experimental_plan": ctx.get("experimental_plan", []),
        "sop": ctx.get("sop", []),
        "process_parameters": ctx.get("process_parameters", []),
        "observations": ctx.get("observations", {"no_anomalies": True, "items": []}),
        "characterization": ctx.get("characterization", []),
        "results": ctx.get("results", {"qualitative": "", "key_data": [], "figures": []}),
        "conclusion": ctx.get("conclusion", ""),
        "next_steps": ctx.get("next_steps", []),
        "original_notes": "",
        "references": list(loop.references),
    }


def _extract_thread_dialogue(loop: "AgentLoop") -> str:
    """从当前线程 history 中提取用户与助手的纯文本对话。
    过滤系统消息、工具调用、工具结果——只保留自然语言往返。"""
    lines: list[str] = []
    for m in loop.history:
        role = m.get("role", "")
        content = (m.get("content") or "").strip()
        if not content:
            continue
        if role == "system":
            continue
        if m.get("tool_calls"):
            continue
        if role == "tool":
            continue
        label = "用户" if role == "user" else "助手"
        lines.append(f"{label}: {content}")
    return "\n".join(lines)


#============================================================================
# 工具日志摘要
# ============================================================================

def _tool_log_summary(name: str, args: dict, result: dict) -> dict:
    """从工具名称、参数和结果中提取关键信息用于日志。"""
    kw = {}
    if name in ("load_reference", "read_experiment"):
        kw["refs"] = args.get("refs", [])
        loaded = ((result.get("loaded", {}) if name == "load_reference"
                   else result.get("experiments", {})) if isinstance(result, dict) else {})
        kw["loaded_count"] = sum(
            1 for v in loaded.values() if isinstance(v, dict) and "error" not in v
        )
    elif name == "search_experiments":
        kw["query"] = args.get("query", "")
        kw["hits"] = len(result.get("candidates", [])) if isinstance(result, dict) else 0
    elif name == "update_schema":
        kw["fields"] = list((args.get("fields") or {}).keys())
    elif name == "generate_record":
        kw["preview_id"] = result.get("id", "")
    elif name == "modify_experiment":
        kw["refs"] = args.get("refs", [])
        kw["fields"] = list((args.get("changes") or {}).keys())
    elif name in ("manage_category", "manage_collection"):
        kw["action"] = args.get("action", "")
        kw["refs"] = args.get("refs", [])
    elif name == "analyze":
        kw["query"] = args.get("query", "")[:100]
    elif name == "list_experiments":
        kw.update({k: v for k, v in args.items() if v})
    elif name == "read_update_log":
        kw["exp_id"] = args.get("exp_id", "")
    if "error" in result:
        kw["error"] = str(result.get("message", result["error"]))[:200]
    return kw


# ============================================================================
# Step 1.2: ToolExecutor
# ============================================================================

class ToolExecutor:
    """注册、校验、执行 LLM 调用的工具"""

    def __init__(self, store, update_log_store=None, favorites_store=None, analysis_store=None,
                 attachment_store=None):
        self.store = store
        self.update_log_store = update_log_store
        self.favorites_store = favorites_store
        self.analysis_store = analysis_store
        self.attachment_store = attachment_store
        self.registry = {
            "load_reference": self._load_reference,
            "search_experiments": self._search_experiments,
            "start_record_thread": self._start_record_thread,
            "update_schema": self._update_schema,
            "generate_record": self._generate_record,
            "read_update_log": self._read_update_log,
            "modify_experiment": self._modify_experiment,
            "manage_archive": self._manage_archive,
            "manage_category": self._manage_category,
            # 兼容已落入历史记录的旧工具名；模型只会看到 manage_category。
            "manage_collection": self._manage_category,
            "read_experiment": self._read_experiment,
            "list_experiments": self._list_experiments,
            "search_chat_history": self._search_chat_history,
            "read_chat_history": self._read_chat_history,
            "list_chat_sessions": self._list_chat_sessions,
            "browse_chat_history": self._browse_chat_history,
            "search_attachments": self._search_attachments,
            "read_attachment": self._read_attachment,
            "manage_attachment": self._manage_attachment,
            "manage_music": self._manage_music,
            "end_thread": self._end_thread,
            "start_analyze_thread": self._start_analyze_thread,
            "select_experiments": self._select_experiments,
            "generate_analysis": self._generate_analysis,
            "read_analysis": self._read_analysis,
        }

    # -- 参数校验入口 --

    def execute(self, name: str, args: dict, loop: "AgentLoop") -> dict:
        """校验参数 → 执行工具。错误以 dict 形式返回，不抛异常。"""
        if name not in self.registry:
            return {"error": "unknown_tool",
                    "message": f"未知工具 '{name}'，可用: {list(self.registry.keys())}"}
        schema = self._tool_schema(name)
        required = schema.get("required", [])
        for key in required:
            if key not in args:
                return {"error": "missing_required",
                        "message": f"缺少必要参数 '{key}'"}
        for key, val in args.items():
            expected = schema["properties"].get(key, {}).get("type")
            if expected == "array" and not isinstance(val, list):
                args[key] = [val]
            elif expected == "string" and isinstance(val, (int, float)):
                args[key] = str(val)
        scope_error = self._child_scope_error(name, args, loop)
        if scope_error:
            return scope_error
        try:
            return self.registry[name](args, loop)
        except Exception as e:
            # 同步层已在冲突后回拉远端真值；把可恢复的并发错误明确交给 Agent，
            # 让它重新读取资源，而不是把它当作普通的工具故障继续猜测。
            from lib.e2ee.syncengine import SyncConflict
            if isinstance(e, SyncConflict):
                return {
                    "error": "remote_revision_conflict",
                    "message": "该资源刚刚在另一设备或窗口被更新。最新数据已同步到本机；请先重新读取，再决定是否修改。",
                }
            return {"error": "execution_failed", "message": str(e)[:300]}

    @staticmethod
    def _child_scope_error(name: str, args: dict, loop: "AgentLoop") -> dict | None:
        """实验编辑子 Agent 只能操作其打开时绑定的那一条实验。"""
        if loop and loop.child.agent_role == "analysis_reviewer":
            anal_id = str(loop.child.exp_id or "").strip()
            if name == "read_analysis" and str(args.get("anal_id") or "").strip() != anal_id:
                return {
                    "error": "child_scope_restricted",
                    "message": f"分析审阅助手只能读取当前报告 {anal_id}。",
                }
        if not loop or loop.child.agent_role != "exp_editor":
            return None
        exp_id = str(loop.child.exp_id or "").strip()
        if not exp_id:
            return {"error": "child_scope_unavailable", "message": "当前编辑会话未绑定实验，不能执行操作。"}
        if name in ("search_experiments", "list_experiments", "search_chat_history",
                    "read_chat_history", "list_chat_sessions", "browse_chat_history"):
            return {"error": "child_scope_restricted", "message": "实验编辑助手只处理当前实验，不能检索其他实验或全局聊天记录。"}
        if name in ("read_experiment", "modify_experiment", "manage_archive"):
            refs = [str(ref).strip() for ref in args.get("refs", []) if str(ref).strip()]
            if refs != [exp_id]:
                return {"error": "child_scope_restricted", "message": f"实验编辑助手只能操作当前实验 {exp_id}。"}
        if name == "read_update_log" and str(args.get("exp_id") or "").strip() != exp_id:
            return {"error": "child_scope_restricted", "message": f"实验编辑助手只能查看当前实验 {exp_id} 的修改历史。"}
        if name == "manage_attachment" and str(args.get("exp_id") or "").strip() != exp_id:
            return {"error": "child_scope_restricted", "message": f"附件只能关联或移出当前实验 {exp_id}。"}
        return None

    def _tool_schema(self, name: str) -> dict:
        """获取工具的 parameters schema"""
        if name == "manage_collection":
            name = "manage_category"
        # 旧会话的历史 tool call 仍可能指向该名称；新模型不会再收到它。
        if name == "load_reference":
            return TOOL_LOAD_REFERENCE["function"]["parameters"]
        for t in TOOLS_OPENAI_FORMAT:
            if t["function"]["name"] == name:
                return t["function"]["parameters"]
        return {}

    # -- start_record_thread --

    def _start_record_thread(self, args: dict, loop: "AgentLoop") -> dict:
        """LLM 判断要开始记录时调用，在当前 user 消息之后插入线程开始标记。"""
        if not loop.thread_store:
            return {"error": "no_thread_store", "message": "线程存储未配置"}
        if loop.thread.id:
            if loop.thread.type == "analyze":
                loop._append_history({"role": "system",
                    "content": f"[系统内部] thread_end id={loop.thread.id}"})
                loop.thread_store.set_active_thread(None)
                loop.thread.id = None
                loop.thread.type = None
            else:
                return {"status": "already_started", "thread_id": loop.thread.id}
        thread_id = loop.thread_store.next_id()
        loop.thread.id = thread_id
        loop.thread.type = "record"
        loop.thread_store.set_active_thread(thread_id)
        loop._enter_record_mode()
        begin = {"role": "system", "content": f"[系统内部] thread_begin id={thread_id} type=record"}
        pos = loop.thread.current_turn_user_idx + 1
        loop._insert_history(pos, begin)
        guidance = {"role": "system", "content": "你正在记录一条新实验。优先收集材料、步骤、参数、结果。追问缺失的关键字段。目标：generate_record。"}
        loop._insert_history(pos + 1, guidance)
        loop.thread_store.create("record", [begin, guidance])
        log = get_logger()
        if log:
            log.operation("thread_start", agent="parent", thread=thread_id, type="record")
        return {"status": "started", "thread_id": thread_id}

    # -- end_thread --

    def _end_thread(self, args: dict, loop: "AgentLoop") -> dict:
        """结束当前线程（record 或 analyze），归档并回到自由模式。"""
        if not loop.thread.id:
            return {"status": "no_active_thread",
                    "message": "当前没有活跃线程。"}
        tid = loop.thread.id
        loop._maybe_inject_thread_end("")
        return {"status": "ended", "thread_id": tid,
                "message": f"线程 {tid} 已结束，回到自由模式。"}

    # -- start_analyze_thread --

    def _start_analyze_thread(self, args: dict, loop: "AgentLoop") -> dict:
        """开启跨实验分析线程。与 start_record_thread 对称。"""
        if not loop.thread_store:
            return {"error": "no_thread_store", "message": "线程存储未配置"}
        if loop.thread.id:
            if loop.thread.type == "record":
                return {"error": "in_record_thread",
                        "message": "当前在 record 线程中。如需分析，请在 record 线程中使用 analyze 工具，或结束 record 线程后再开启 analyze 线程。"}
            if loop.thread.type == "analyze":
                return {"status": "already_started", "thread_id": loop.thread.id}
        thread_id = loop.thread_store.next_id()
        loop.thread.id = thread_id
        loop.thread.type = "analyze"
        loop.thread_store.set_active_thread(thread_id)
        begin = {"role": "system", "content": f"[系统内部] thread_begin id={thread_id} type=analyze"}
        pos = loop.thread.current_turn_user_idx + 1
        loop._insert_history(pos, begin)
        guidance = loop._build_thread_guidance("analyze")
        loop._insert_history(pos + 1, guidance)
        loop.thread_store.create("analyze", [begin, guidance])
        log = get_logger()
        if log:
            log.operation("thread_start", agent="parent", thread=thread_id, type="analyze")
        return {"status": "started", "thread_id": thread_id}

    # -- select_experiments --

    def _select_experiments(self, args: dict, loop: "AgentLoop") -> dict:
        """返回选择面板数据，由前端渲染为实验勾选卡片。"""
        return {
            "display": "selector",
            "pause": True,
            "title": args.get("title", "选择实验"),
            "items": args.get("candidates", []),
            "preselected": args.get("preselected", []),
        }

    # -- generate_analysis --

    def _generate_analysis(self, args: dict, loop: "AgentLoop") -> dict:
        """执行分析 → 写 AnalysisStore → 自动结束线程 → 返回标题+摘要。"""
        query = args["query"]
        refs = []
        if loop.thread_store and loop.thread.id:
            thread = loop.thread_store.load(loop.thread.id)
            if thread and thread.get("type") == "analyze":
                refs = list(thread.get("selected_exps") or [])
        if len(refs) < 2:
            return {"error": "analysis_selection_required",
                    "message": "请先在实验选择面板确认至少 2 个实验。"}
        try:
            service = loop.analysis_svc
            if service is None:
                if not self.analysis_store:
                    return {"error": "no_analysis_store", "message": "分析存储未配置"}
                from lib.services.analysis import AnalysisService
                service = AnalysisService(
                    self.store, self.analysis_store, loop.llm,
                    update_log_repo=self.update_log_store,
                    attachment_store=self.attachment_store,
                )
            result = service.run_analysis(query, refs)
            anal_id = result["anal_id"]
            topic = result.get("topic") or query
            excerpt = result["analysis"][:200]
            tid = loop.thread.id
            if tid and not loop.child.is_child:
                loop._maybe_inject_thread_end(anal_id)
            return {
                "display": "analysis_done",
                "anal_id": anal_id,
                # title 保留给历史渲染兼容；新卡片优先使用 topic。
                "title": topic,
                "topic": topic,
                "summary": excerpt,
                "refs": refs,
            }
        except (APITimeoutError, TimeoutError):
            timeout_seconds = (loop.analysis_svc.request_timeout_seconds()
                               if loop.analysis_svc else 8 * 60)
            return {
                "error": "analysis_timeout",
                "message": f"分析报告在 {timeout_seconds // 60} 分钟上限内未完成；本次未自动重试。",
            }
        except Exception as e:
            return {"error": "analysis_failed", "message": str(e)[:300]}

    def _read_analysis(self, args: dict, loop: "AgentLoop") -> dict:
        """读取已归档分析报告；报告本身保持只读。"""
        if not self.analysis_store:
            return {"error": "no_analysis_store", "message": "分析存储未配置"}
        anal_id = str(args.get("anal_id") or "").strip()
        if not anal_id:
            return {"error": "missing_anal_id", "message": "请提供分析编号"}
        analysis = self.analysis_store.load(anal_id)
        if not analysis:
            return {"error": "not_found", "message": f"分析报告 {anal_id} 不存在"}
        return {
            "status": "ok",
            "analysis": {
                "id": analysis.get("id", anal_id),
                "timestamp": analysis.get("timestamp", ""),
                "question": analysis.get("question", ""),
                "selected_ids": analysis.get("selected_ids", []),
                "content": analysis.get("analysis", ""),
            },
        }

    # -- generate_record --

    def _generate_record(self, args: dict, loop: "AgentLoop") -> dict:
        # 子Agent 不允许 generate_record → 使用 modify_experiment 直接修改
        if loop.child.is_child:
            return {"error": "use_modify_experiment",
                    "message": "子Agent请使用 modify_experiment 工具直接修改实验字段。修改会自动保存。"}
        if loop._schema_context is None:
            return {"error": "not_in_record_mode",
                    "message": "generate_record 只在记录实验时可用。"}

        notes = loop._build_notes_from_context()

        # 构建增强 prompt: 四段式 = RAW SCHEMA + DIALOGUE + NOTES + REFERENCES
        import json as _json
        prompt_parts = []

        # 段 1: 原始 Schema JSON —— 让提取 LLM 精确知道哪些字段已填、哪些缺失
        raw_schema = _json.dumps(
            loop._schema_context, ensure_ascii=False, indent=2)
        prompt_parts.append(
            "---RAW SCHEMA (current field values, empty means unfilled)---\n"
            + raw_schema)

        # 段 2: 线程纯文本对话 —— 保留用户原始措辞（近似值、事后补充等细节）
        dialogue = _extract_thread_dialogue(loop)
        if dialogue:
            prompt_parts.append(
                "---DIALOGUE (original conversation for nuance)---\n"
                + dialogue)

        # 段 3: 自然语言实验描述 —— 帮助 LLM 理解语义连贯性
        prompt_parts.append(
            "---NOTES TEXT (structured summary of the experiment)---\n"
            + notes)

        # 段 4: 已加载引用实验的结构化摘要 —— 帮助校验和恢复漏掉的字段
        if loop.references:
            ref_parts = []
            for ref_id in loop.references:
                ref_exp = loop.store.load(ref_id)
                if ref_exp:
                    ref_parts.append(
                        _json.dumps(self._summarize_exp(ref_exp),
                                    ensure_ascii=False, indent=2))
            if ref_parts:
                prompt_parts.append(
                    "---REFERENCES (loaded experiments for comparison, "
                    "do NOT copy conclusion/results directly)---\n"
                    + "\n".join(ref_parts))

        enhanced_notes = "\n\n".join(prompt_parts)

        try:
            from lib.parser import parse_notes
            result = parse_notes(enhanced_notes, loop.llm)
            result["original_notes"] = notes
            # 子 Agent: 使用现有 EXP ID（修改已有实验）
            if loop.child.is_child and loop.child.exp_id:
                result["id"] = loop.child.exp_id
            else:
                result["id"] = loop.store.next_id()
            result["references"] = list(loop.references)
            loop._generated_preview = result
            loop._generated_notes = notes
            card_summary = result.get("title", "")
            conclusion = (result.get("conclusion") or "").strip()
            if conclusion:
                card_summary += " — " + conclusion[:40]
            return {"status": "generated", "pause": True,
                    "display": "record_generated",
                    "exp_id": result["id"],
                    "summary": card_summary,
                    "response_type": "generate", "include_state": True,
                    "id": result["id"],
                    "title": result.get("title", ""),
                    "fields_count": sum(1 for v in result.values() if v)}
        except Exception:
            preview = _build_preview(loop)
            # 子 Agent 回退也使用现有 EXP ID
            if loop.child.is_child and loop.child.exp_id:
                preview["id"] = loop.child.exp_id
            loop._generated_preview = preview
            loop._generated_notes = notes
            card_summary = preview.get("title", "")
            conclusion = (preview.get("conclusion") or "").strip()
            if conclusion:
                card_summary += " — " + conclusion[:40]
            return {"status": "generated", "pause": True,
                    "display": "record_generated",
                    "exp_id": preview["id"],
                    "summary": card_summary,
                    "response_type": "generate", "include_state": True,
                    "id": preview["id"],
                    "title": preview.get("title", ""),
                    "note": "LLM 提取失败，使用了确定性回退，部分字段可能需手动补全"}

    # -- read_update_log --

    def _read_update_log(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.update_log_store:
            return {"error": "no_update_log_store", "message": "更新日志存储未配置"}
        exp_id = args["exp_id"]
        limit = args.get("limit", 5)
        entries = self.update_log_store.list_recent(exp_id, limit=limit)
        return {"status": "ok", "exp_id": exp_id, "entries": entries}

    # -- modify_experiment --

    def _modify_experiment(self, args: dict, loop: "AgentLoop") -> dict:
        refs = args.get("refs", [])
        changes = args.get("changes", {})
        if not refs:
            return {"error": "no_refs", "message": "请指定要修改的实验编号"}
        if not changes:
            return {"error": "no_changes", "message": "请指定要修改的字段"}

        results = {}
        display_changes = []
        for ref in refs:
            exp = self.store.load(ref)
            if not exp:
                results[ref] = {"error": "not_found", "message": f"实验 {ref} 不存在"}
                continue
            # 读磁盘旧值
            old_exp = deepcopy(exp)
            expected_revision = args.get("expected_revision")
            if expected_revision is None:
                expected_revision = int(exp.get("revision", 0) or 0)
            else:
                try:
                    expected_revision = int(expected_revision)
                except (TypeError, ValueError):
                    results[ref] = {"error": "invalid_revision", "message": "revision 必须是整数"}
                    continue
            # 应用 changes
            for key, value in changes.items():
                if key in ("materials", "equipment", "experimental_plan",
                          "process_parameters", "characterization"):
                    exp[key] = value  # 完整替换
                elif key in ("results", "observations"):
                    if isinstance(value, dict):
                        exp.setdefault(key, {}).update(value)
                elif key == "tags":
                    exp[key] = list(value)
                elif key == "sop" or key == "next_steps":
                    exp[key] = list(value)
                else:
                    exp[key] = value
            # 先计算 diff；只有条件保存成功后才写日志，避免冲突请求留下伪记录。
            from lib.services.experiment import compute_experiment_diff
            entries = compute_experiment_diff(old_exp, exp)
            # 条件保存：另一窗口在本次读取之后已更新时，不能静默覆盖。
            if hasattr(self.store, "save_if_revision"):
                saved = self.store.save_if_revision(exp, expected_revision)
                if not saved.get("ok"):
                    results[ref] = {
                        "error": saved.get("error", "save_failed"),
                        "message": "实验已在另一窗口更新，请重新读取后再修改。",
                        "revision": saved.get("revision"),
                    }
                    continue
            else:
                self.store.save(exp)
            if entries and self.update_log_store:
                self.update_log_store.append(
                    exp_id=ref, source="parent_agent",
                    changes=entries,
                    thread_id=loop.thread.id,
                    context={"summary": f"修改了 {len(entries)} 个字段"},
                )
            # 注入过期标记到 history
            loop._append_history({
                "role": "system",
                "content": f"{ref} 已被修改。此前关于 {ref} 的对话陈述可能已过时。获取当前数据请使用 read_experiment。"
            })
            results[ref] = {
                "status": "modified",
                "display": "diff",
                "changes": entries,
            }
            display_changes.extend({**entry, "exp_id": ref} for entry in entries)
        result = {"modified": results}
        if display_changes:
            result.update({"display": "diff", "changes": display_changes})
        return result

    # -- manage_archive --

    def _manage_archive(self, args: dict, loop: "AgentLoop") -> dict:
        action = args.get("action")
        refs = [str(ref).strip() for ref in args.get("refs", []) if str(ref).strip()]
        if action not in ("archive", "restore") or not refs:
            return {"error": "invalid_archive_request", "message": "请提供归档动作和实验编号"}
        archived = action == "archive"
        results = {}
        for ref in refs:
            expected_revision = args.get("expected_revision")
            try:
                expected_revision = int(expected_revision) if expected_revision is not None else None
            except (TypeError, ValueError):
                results[ref] = {"error": "invalid_revision", "message": "revision 必须是整数"}
                continue
            if not hasattr(self.store, "set_archived"):
                results[ref] = {"error": "not_supported", "message": "当前存储不支持归档"}
                continue
            saved = self.store.set_archived(ref, archived, expected_revision)
            if not saved.get("ok"):
                results[ref] = {
                    "error": saved.get("error", "save_failed"),
                    "message": "实验已在另一窗口更新，请重新读取后再操作。",
                    "revision": saved.get("revision"),
                }
                continue
            if self.update_log_store:
                self.update_log_store.append(
                    exp_id=ref, source="parent_agent",
                    changes=[{"path": "archived", "field": "归档状态",
                              "old": "正常" if archived else "已归档",
                              "new": "已归档" if archived else "已恢复到实验列表"}],
                    thread_id=loop.thread.id,
                    context={"summary": f"实验已{'归档' if archived else '恢复'}"},
                )
            loop._append_history({
                "role": "system",
                "content": f"{ref} 已{'归档' if archived else '恢复到实验列表'}；读取时以当前磁盘状态为准。",
            })
            results[ref] = {"status": "archived" if archived else "restored",
                            "revision": saved.get("revision")}
        changed = [ref for ref, item in results.items() if item.get("status")]
        return {
            "status": "ok" if len(changed) == len(refs) else "partial",
            "display": "toast",
            "action": action,
            "changed": bool(changed),
            "refs": changed,
            "results": results,
            "message": f"已{'归档' if archived else '恢复'} {len(changed)} 条实验记录",
        }

    # -- manage_category --

    def _manage_category(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.favorites_store:
            return {"error": "no_category_store", "message": "分类存储未配置"}
        action = {"favorite": "add", "unfavorite": "remove"}.get(args["action"], args["action"])
        category = str(args.get("category") or args.get("collection") or "").strip()
        name = str(args.get("name") or "").strip()
        refs = [str(ref).strip() for ref in args.get("refs", []) if str(ref).strip()]

        if action == "list":
            collections = self.favorites_store.get_collections()
            meta = self.favorites_store.get_collection_meta()
            category_pinned = self.favorites_store.get_category_pinned()
            items = []
            for item_name, ids in collections.items():
                if category and item_name != category:
                    continue
                items.append({
                    "name": item_name,
                    "description": (meta.get(item_name) or {}).get("description", ""),
                    "experiments": [
                        {"id": exp_id, "title": (self.store.load(exp_id) or {}).get("title", "")}
                        for exp_id in ids
                    ],
                    "pinned": category_pinned.get(item_name, []),
                })
            if category and not items:
                return {"error": "category_not_found", "message": f"分类不存在：{category}"}
            return {"status": "ok", "categories": items,
                    "global_pinned": self.favorites_store.get_pinned()}

        if action in ("create", "rename") and not name:
            return {"error": "missing_name", "message": "请提供分类名称"}
        if action in ("rename", "delete", "add", "remove") and not category:
            return {"error": "missing_category", "message": "请提供分类名称"}
        if action in ("add", "remove", "pin", "unpin") and not refs:
            return {"error": "missing_refs", "message": "请提供实验编号"}

        if action == "create":
            result = self.favorites_store.create_collection(name)
            return self._category_mutation_result(action, {name: result})
        if action == "rename":
            result = self.favorites_store.update_collection(category, new_name=name)
            return self._category_mutation_result(action, {category: result})
        if action == "delete":
            result = self.favorites_store.delete_collection(category)
            return self._category_mutation_result(action, {category: result})

        results = {}
        for ref in refs:
            if not self.store.load(ref):
                results[ref] = {"ok": False, "error": "experiment_not_found"}
                continue
            if action in ("add", "remove"):
                if category not in self.favorites_store.get_collections():
                    results[ref] = {"ok": False, "error": "category_not_found"}
                else:
                    results[ref] = self.favorites_store.set_favorite(ref, category, action == "add")
            elif category:
                current = self.favorites_store.is_category_pinned(ref, category)
                if current == (action == "pin"):
                    results[ref] = {"ok": True, "pinned": current, "category": category}
                else:
                    results[ref] = self.favorites_store.toggle_category_pin(ref, category)
            else:
                results[ref] = self.favorites_store.set_pin(ref, action == "pin")
        return self._category_mutation_result(action, results)

    @staticmethod
    def _category_mutation_result(action: str, results: dict) -> dict:
        succeeded = sum(1 for result in results.values() if result.get("ok"))
        status = "ok" if succeeded == len(results) else "partial"
        return {"status": status, "display": "toast", "changed": succeeded > 0,
                "message": f"分类已{('更新' if status == 'ok' else '部分更新')}（{succeeded}/{len(results)}）",
                "action": action, "results": results}

    # -- read_experiment --

    def _read_experiment(self, args: dict, loop: "AgentLoop") -> dict:
        refs = args.get("refs", [])
        requested = args.get("fields") or []
        allowed = {"id", "revision", "created_at", "updated_at", *FIELD_DEFAULTS.keys()}
        fields = [str(field) for field in requested if str(field) in allowed]
        if requested and not fields:
            return {"error": "invalid_fields", "message": "未请求到可读取的实验字段"}
        result = {}
        registered_references = []
        for ref in refs:
            exp_id = str(ref).strip()
            exp = self.store.load(exp_id)
            if not exp:
                result[exp_id] = {"error": "not_found"}
                continue
            selected = fields or ["id", "revision", "created_at", "updated_at", *FIELD_DEFAULTS.keys()]
            data = {field: deepcopy(exp.get(field)) for field in selected if field in exp}
            if args.get("include_updates") and self.update_log_store:
                data["recent_updates"] = self.update_log_store.list_recent(exp_id, limit=3)
            result[exp_id] = data
            if args.get("as_reference") and exp_id not in loop.references:
                loop.references.append(exp_id)
                registered_references.append(exp_id)
            if args.get("as_reference") and loop.experiment_type == "other":
                for tag in exp.get("tags") or []:
                    if tag in ("photocatalysis", "hydrothermal", "sol-gel",
                               "spin-coating", "ball-milling", "electrochemistry",
                               "xrd", "perovskite-solar"):
                        loop.experiment_type = tag
                        break
        response = {"status": "ok", "experiments": result}
        if args.get("as_reference"):
            response["registered_references"] = registered_references
        return response

    # -- chat_history --

    def _search_chat_history(self, args: dict, loop: "AgentLoop") -> dict:
        if not loop.thread_store or not hasattr(loop.thread_store, "search_chat_history"):
            return {"error": "chat_history_unavailable", "message": "聊天记录检索尚未配置"}
        query = str(args.get("query") or "").strip()
        if not query:
            return {"error": "empty_query", "message": "请提供要查找的关键词"}
        try:
            limit = max(1, min(int(args.get("limit") or 10), 20))
        except (TypeError, ValueError):
            limit = 10
        matches = loop.thread_store.search_chat_history(
            query, date_from=args.get("date_from") or None,
            date_to=args.get("date_to") or None, limit=limit,
        )
        return {"status": "ok", "query": query, "count": len(matches), "matches": matches}

    def _read_chat_history(self, args: dict, loop: "AgentLoop") -> dict:
        if not loop.thread_store or not hasattr(loop.thread_store, "read_chat_history_context"):
            return {"error": "chat_history_unavailable", "message": "聊天记录检索尚未配置"}
        try:
            sequence = int(args["sequence"])
            before = int(args.get("before") or 3)
            after = int(args.get("after") or 3)
        except (TypeError, ValueError):
            return {"error": "invalid_reference", "message": "聊天记录定位参数无效"}
        context = loop.thread_store.read_chat_history_context(
            str(args["session_id"]), sequence, before, after,
        )
        if not context:
            return {"error": "not_found", "message": "未找到该聊天记录"}
        return {"status": "ok", **context}

    def _list_chat_sessions(self, args: dict, loop: "AgentLoop") -> dict:
        if not loop.thread_store or not hasattr(loop.thread_store, "list_chat_sessions"):
            return {"error": "chat_history_unavailable", "message": "聊天记录检索尚未配置"}
        try:
            limit = max(1, min(int(args.get("limit") or 20), 100))
        except (TypeError, ValueError):
            limit = 20
        return {"status": "ok", "sessions": loop.thread_store.list_chat_sessions(limit)}

    def _browse_chat_history(self, args: dict, loop: "AgentLoop") -> dict:
        if not loop.thread_store or not hasattr(loop.thread_store, "browse_chat_history"):
            return {"error": "chat_history_unavailable", "message": "聊天记录检索尚未配置"}
        try:
            before = args.get("before_sequence")
            before = int(before) if before is not None else None
            limit = max(1, min(int(args.get("limit") or 20), 100))
        except (TypeError, ValueError):
            return {"error": "invalid_page", "message": "聊天记录分页参数无效"}
        return {"status": "ok", **loop.thread_store.browse_chat_history(
            str(args.get("session_id") or ""), before, limit
        )}

    # -- attachments --

    def _attachment_links(self) -> dict[str, list[dict[str, Any]]]:
        links: dict[str, list[dict[str, Any]]] = {}
        for experiment in self.store.list_all_full():
            for attachment in experiment.get("attachments") or []:
                if not isinstance(attachment, dict) or not attachment.get("sha256"):
                    continue
                links.setdefault(attachment["sha256"], []).append({
                    "exp_id": experiment.get("id", ""),
                    "experiment": experiment.get("title", ""),
                    "title": attachment.get("title") or attachment.get("caption") or attachment.get("name", ""),
                    "description": attachment.get("description", ""),
                })
        return links

    def _child_visible_attachment_ids(self, loop: "AgentLoop") -> set[str]:
        """当前实验附件，加上本次编辑对话里由用户上传的附件。"""
        if loop.child.agent_role != "exp_editor":
            return set()
        experiment = self.store.load(loop.child.exp_id) or {}
        visible = {
            str(item.get("sha256")) for item in (experiment.get("attachments") or [])
            if isinstance(item, dict) and item.get("sha256")
        }
        for message in loop.history:
            for attachment in message.get("attachments") or []:
                if isinstance(attachment, dict) and attachment.get("sha256"):
                    visible.add(str(attachment["sha256"]))
        return visible

    def _search_attachments(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.attachment_store:
            return {"error": "attachment_store_unavailable", "message": "附件存储尚未配置"}
        query = str(args.get("query") or "").strip().casefold()
        try:
            limit = max(1, min(int(args.get("limit") or 20), 50))
        except (TypeError, ValueError):
            limit = 20
        links = self._attachment_links()
        visible_ids = self._child_visible_attachment_ids(loop) if loop else set()
        results = []
        for meta in self.attachment_store.list_all():
            if loop and loop.child.agent_role == "exp_editor" and meta.get("sha256") not in visible_ids:
                continue
            locations = links.get(meta["sha256"], [])
            searchable = " ".join([
                meta.get("name", ""),
                *[f"{item.get('title', '')} {item.get('description', '')}" for item in locations],
            ]).casefold()
            if query and query not in searchable:
                continue
            results.append({
                "sha256": meta["sha256"], "name": meta.get("name", ""),
                "mime": meta.get("mime", ""), "size": meta.get("size", 0),
                "linked_experiments": locations,
            })
            if len(results) >= limit:
                break
        return {"status": "ok", "query": query, "count": len(results), "attachments": results}

    def _read_attachment(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.attachment_store:
            return {"error": "attachment_store_unavailable", "message": "附件存储尚未配置"}
        sha256 = str(args.get("sha256") or "").strip()
        if loop and loop.child.agent_role == "exp_editor" and sha256 not in self._child_visible_attachment_ids(loop):
            return {"error": "child_scope_restricted", "message": "实验编辑助手只能读取当前实验或本次对话上传的附件。"}
        meta = self.attachment_store.meta(sha256)
        content = self.attachment_store.get(sha256)
        if not meta or content is None:
            return {"error": "attachment_not_found", "message": "附件不存在或尚未同步到本机"}
        mime = meta.get("mime", "")
        name = meta.get("name", "")
        suffix = Path(name).suffix.lower()
        is_xlsx = suffix == ".xlsx" or mime in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        default_max_chars = 4000 if is_xlsx else 12000
        try:
            max_chars = max(1, min(int(args.get("max_chars") or default_max_chars), 30000))
        except (TypeError, ValueError):
            max_chars = default_max_chars
        text = ""
        extraction = "text"
        try:
            if mime.startswith("text/") or suffix in (".csv", ".txt", ".tsv"):
                text = content.decode("utf-8-sig", errors="replace")
            elif is_xlsx:
                from io import BytesIO
                from openpyxl import load_workbook
                book = load_workbook(BytesIO(content), read_only=True, data_only=True)
                requested_sheet = str(args.get("sheet") or "").strip()
                if requested_sheet and requested_sheet not in book.sheetnames:
                    book.close()
                    return {"error": "attachment_sheet_not_found", "name": name,
                            "message": f"工作表不存在：{requested_sheet}"}
                try:
                    start_row = max(1, int(args.get("start_row") or 1))
                except (TypeError, ValueError):
                    start_row = 1

                sheet_count = len(book.worksheets)
                selected_sheets = [book[requested_sheet]] if requested_sheet else book.worksheets
                sheets = []
                for sheet in selected_sheets:
                    total_rows = sheet.max_row or 0
                    total_columns = sheet.max_column or 0
                    max_column = min(total_columns, 30)
                    end_row = min(total_rows, start_row + 99)
                    rows = []
                    chars_used = 0
                    char_truncated = False
                    if max_column and start_row <= end_row:
                        for row in sheet.iter_rows(
                            min_row=start_row, max_row=end_row, max_col=max_column, values_only=True,
                        ):
                            row_text = "\t".join("" if value is None else str(value) for value in row)
                            separator = 1 if rows else 0
                            if chars_used + separator + len(row_text) > max_chars:
                                char_truncated = True
                                if not rows:
                                    rows.append(row_text[:max_chars])
                                break
                            rows.append(row_text)
                            chars_used += separator + len(row_text)

                    shown_rows = len(rows)
                    truncated_by_rows = start_row + shown_rows - 1 < total_rows
                    truncation_reasons = []
                    if char_truncated:
                        truncation_reasons.append("chars")
                    if truncated_by_rows and not char_truncated:
                        truncation_reasons.append("rows")
                    if total_columns > max_column:
                        truncation_reasons.append("columns")
                    sheets.append({
                        "name": sheet.title,
                        "rows": total_rows,
                        "columns": total_columns,
                        "start_row": start_row,
                        "shown_rows": shown_rows,
                        "shown_columns": max_column,
                        "content": "\n".join(rows),
                        "truncated": bool(truncation_reasons),
                        "truncation_reasons": truncation_reasons,
                        "next_start_row": start_row + shown_rows if truncated_by_rows else None,
                    })
                book.close()
                return {"status": "ok", "name": name, "mime": mime, "size": len(content),
                        "extraction": "xlsx", "max_chars_per_sheet": max_chars,
                        "sheet_count": sheet_count, "sheets": sheets}
            elif suffix == ".pdf" or mime == "application/pdf":
                from io import BytesIO
                from pypdf import PdfReader
                reader = PdfReader(BytesIO(content))
                text = "\n\n".join((page.extract_text() or "") for page in reader.pages[:30])
                extraction = "pdf"
            elif mime.startswith("image/") or suffix in (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"):
                llm = getattr(loop, "llm", None)
                if llm is None or not hasattr(llm, "describe_image"):
                    return {
                        "status": "vision_unavailable", "name": name, "mime": mime,
                        "size": meta.get("size", 0), "extraction": "vision",
                        "message": "当前对话模型未提供图片阅读能力。请在设置中选择支持视觉的模型后重试。",
                    }
                try:
                    text = llm.describe_image(content, mime, name)
                except Exception as exc:
                    return {
                        "status": "vision_unavailable", "name": name, "mime": mime,
                        "size": meta.get("size", 0), "extraction": "vision",
                        "message": f"当前模型无法读取此图片：{str(exc)[:160]}",
                    }
                extraction = "vision"
            else:
                return {"status": "metadata_only", "name": name, "mime": mime,
                        "size": meta.get("size", 0),
                        "message": "此文件类型暂不支持读取正文。"}
        except (ImportError, OSError, ValueError, RuntimeError) as exc:
            return {"status": "metadata_only", "name": name, "mime": mime,
                    "size": meta.get("size", 0),
                    "message": f"无法提取附件正文：{str(exc)[:160]}"}
        if not text.strip():
            return {"status": "metadata_only", "name": name, "mime": mime,
                    "size": meta.get("size", 0), "message": "文件没有可提取的文本内容。"}
        return {"status": "ok", "name": name, "mime": mime, "size": len(content),
                "content": text[:max_chars], "truncated": len(text) > max_chars,
                "extraction": extraction}

    def _manage_attachment(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.attachment_store:
            return {"error": "attachment_store_unavailable", "message": "附件存储尚未配置"}
        action = args.get("action")
        sha256 = str(args.get("sha256") or "").strip()
        exp_id = str(args.get("exp_id") or "").strip()
        if loop and loop.child.agent_role == "exp_editor" and sha256 not in self._child_visible_attachment_ids(loop):
            return {"error": "child_scope_restricted", "message": "实验编辑助手只能关联当前实验或本次对话上传的附件。"}
        meta = self.attachment_store.meta(sha256)
        experiment = self.store.load(exp_id)
        if not meta:
            return {"error": "attachment_not_found", "message": "附件不存在"}
        if not experiment:
            return {"error": "experiment_not_found", "message": "目标实验不存在"}
        links = [item for item in (experiment.get("attachments") or [])
                 if isinstance(item, dict) and item.get("sha256")]
        if action == "unlink":
            kept = [item for item in links if item["sha256"] != sha256]
            if len(kept) == len(links):
                return {"error": "link_not_found", "message": "该附件未关联到此实验"}
            experiment["attachments"] = kept
            if hasattr(self.store, "save_if_revision"):
                saved = self.store.save_if_revision(experiment, int(experiment.get("revision", 0) or 0))
                if not saved.get("ok"):
                    return {"error": "revision_conflict", "message": "实验已在另一窗口更新，请重新读取后再操作。"}
            else:
                self.store.save(experiment)
            return {"status": "ok", "action": "unlinked", "sha256": sha256, "exp_id": exp_id}
        if action != "link":
            return {"error": "invalid_action", "message": "action 必须是 link 或 unlink"}
        link = {
            "sha256": sha256, "name": meta.get("name", ""), "mime": meta.get("mime", ""),
            "size": meta.get("size", 0), "title": str(args.get("title") or meta.get("name") or "附件")[:200],
            "description": str(args.get("description") or "")[:1000],
            "created_at": meta.get("created_at", ""),
        }
        existing = next((item for item in links if item["sha256"] == sha256), None)
        if existing:
            existing.update(link)
        else:
            links.append(link)
        experiment["attachments"] = links
        if hasattr(self.store, "save_if_revision"):
            saved = self.store.save_if_revision(experiment, int(experiment.get("revision", 0) or 0))
            if not saved.get("ok"):
                return {"error": "revision_conflict", "message": "实验已在另一窗口更新，请重新读取后再操作。"}
        else:
            self.store.save(experiment)
        return {"status": "ok", "action": "linked", "sha256": sha256, "exp_id": exp_id,
                "title": link["title"]}

    def _manage_music(self, args: dict, loop: "AgentLoop") -> dict:
        if not self.attachment_store:
            return {"error": "attachment_store_unavailable", "message": "音乐库尚未配置"}
        from lib.music import is_audio_attachment, library_for

        action = str(args.get("action") or "")
        tracks = library_for(self.attachment_store)
        track_map = {track["id"]: track for track in tracks}
        state = self.attachment_store.get_music_playback()
        if action == "status":
            return {"status": "ok", "action": "status", "display": "music_control",
                    "track": track_map.get(str(state.get("track_id") or "")), "current": state,
                    "tracks": [{"id": track["id"], "title": track["title"]} for track in tracks]}
        if action == "add":
            sha256 = str(args.get("sha256") or "").strip()
            meta = self.attachment_store.meta(sha256)
            if not meta:
                return {"error": "attachment_not_found", "message": "找不到该上传附件"}
            if not is_audio_attachment(str(meta.get("name") or ""), str(meta.get("mime") or "")):
                return {"error": "not_audio_attachment", "message": "该附件不是支持的音频文件"}
            added = self.attachment_store.add_music_track(sha256, str(args.get("title") or ""))
            if not added:
                return {"error": "attachment_not_found", "message": "附件内容尚未下载到本机"}
            track = next(item for item in library_for(self.attachment_store)
                         if item["id"] == f"attachment:{sha256}")
            return {"status": "ok", "action": "add", "display": "music_control",
                    "track": track, "message": f"已加入曲库：{track['title']}"}
        if action == "stop":
            track = track_map.get(str(state.get("track_id") or ""))
            state = self.attachment_store.set_music_playback(False)
            return {"status": "ok", "action": "stop", "display": "music_control",
                    "track": track, "current": state}
        if action not in ("play", "next"):
            return {"error": "invalid_action", "message": "action 必须是 status、play、stop、next 或 add"}
        track_id = str(args.get("track_id") or "")
        if action == "next":
            current_id = str(state.get("track_id") or "")
            choices = [track for track in tracks if track["id"] != current_id] or tracks
            if choices:
                import random
                track_id = random.choice(choices)["id"]
        if not track_id:
            track_id = str(state.get("track_id") or (tracks[0]["id"] if tracks else ""))
        track = track_map.get(track_id)
        if not track:
            return {"error": "track_not_found", "message": "曲目不存在；请先用 status 查看曲库"}
        state = self.attachment_store.set_music_playback(True, track_id)
        return {"status": "ok", "action": action, "display": "music_control",
                "track": track, "current": state}

    # -- list_experiments --

    def _list_experiments(self, args: dict, loop: "AgentLoop") -> dict:
        all_exps = self.store.list_all_full(include_archived=bool(args.get("include_archived")))
        filtered = []
        status = args.get("status")
        tags = args.get("tags", [])
        experimenter = args.get("experimenter")
        since = args.get("since")

        for exp in all_exps:
            if status and exp.get("status") != status:
                continue
            if tags:
                exp_tags = [t.lower() for t in exp.get("tags", [])]
                if not any(t.lower() in exp_tags for t in tags):
                    continue
            if experimenter and exp.get("experimenter") != experimenter:
                continue
            if since and exp.get("date", "") < since:
                continue
            filtered.append({
                "id": exp.get("id"),
                "title": exp.get("title", ""),
                "date": exp.get("date", ""),
                "status": exp.get("status", ""),
                "tags": exp.get("tags", []),
            })
        return {
            "display": "list",
            "experiments": filtered[:20],
            "count": len(filtered),
        }

    # -- 旧工具兼容：新模型应使用 read_experiment(as_reference=true) --

    def _load_reference(self, args: dict, loop: "AgentLoop") -> dict:
        """兼容旧名称；返回完整读取结果，且登记引用。"""
        response = self._read_experiment({**args, "as_reference": True}, loop)
        return {
            "loaded": response["experiments"],
            "registered_references": response.get("registered_references", []),
        }

    def _summarize_exp(self, exp: dict) -> dict:
        """引用快照保留全部结构化字段，不能丢失实验员、附件等继承信息。"""
        result = {
            field: deepcopy(exp.get(field))
            for field in ("id", "revision", "created_at", "updated_at", *FIELD_DEFAULTS.keys())
            if field in exp
        }
        # 追加最近更新日志摘要
        if self.update_log_store:
            try:
                recent = self.update_log_store.list_recent(exp.get("id", ""), limit=3)
                if recent:
                    result["_recent_updates"] = [
                        {"timestamp": r.get("timestamp", ""),
                         "source": r.get("source", ""),
                         "summary": r.get("context", {}).get("summary", ""),
                         "changed_fields": [c.get("field", "") for c in r.get("changes", [])]}
                        for r in recent
                    ]
            except Exception:
                pass
        return result

    # -- Step 1.6: search_experiments --

    def _search_experiments(self, args: dict, loop: "AgentLoop") -> dict:
        query = args.get("query", "").strip()
        if not query or len(query) < 2:
            return {"candidates": []}

        # 第一步：关键词粗筛
        include_archived = bool(args.get("include_archived"))
        keyword_results = self._fuzzy_search(query, loop, include_archived)

        # 第二步：FTS 全文搜索（SQLite 引擎，毫秒级）
        best_kw = keyword_results[0]["score"] if keyword_results else 0.0
        if hasattr(self.store, "search") and best_kw < 0.5:
            fts_results = self.store.search(query, include_archived=include_archived)
            if fts_results:
                candidates = [
                    {"id": r["id"], "title": r.get("title", ""), "date": r.get("date", ""),
                     "tags": r.get("tags", []), "score": r.get("_score", 0.5)} for r in fts_results
                ]
                # 合并：FTS 结果优先，关键词结果补充
                merged = {c["id"]: c for c in candidates}
                for kw in keyword_results:
                    if kw["id"] not in merged:
                        merged[kw["id"]] = kw
                return {"candidates": sorted(merged.values(), key=lambda x: -x["score"])[:10]}

        # 如果是纯 ID/编号查询，关键词就够了
        if re.match(r'^[\w-]*\d[\w-]*$', query) or is_experiment_id(query):
            return {"candidates": keyword_results[:5]}

        # 第二步：自然语言查询 → LLM 语义搜索
        if not keyword_results or keyword_results[0]["score"] < 0.3:
            try:
                llm_results = self._llm_semantic_search(query, loop, include_archived)
                if llm_results:
                    return {"candidates": llm_results[:5]}
            except Exception:
                pass

        return {"candidates": keyword_results[:5]}

    def _llm_semantic_search(self, query: str, loop: "AgentLoop", include_archived: bool = False) -> list[dict]:
        """LLM 语义搜索：独立 API 调用，不污染 Agent 上下文。处理自然语言如'上周一的''老张做的''失败的那个'。"""
        all_exps = loop.store.list_all_full(include_archived=include_archived)
        if not all_exps:
            return []

        # 构造极简摘要（每实验 1-2 行，控制 token 消耗）
        lines = []
        for e in all_exps:
            exp_id = e.get("id", "")
            title = (e.get("title") or "(无标题)")[:40]
            date = e.get("date") or ""
            experimenter = e.get("experimenter") or "佚名"
            status = e.get("status", "")
            status_cn = {"planned": "计划中", "running": "进行中", "done": "已完成",
                         "failed": "失败", "repeated": "重复"}.get(status, status)
            conclusion = (e.get("conclusion") or "")[:40]
            tags = ", ".join(e.get("tags", [])[:4])
            lines.append(
                f"{exp_id} | {title} | {date} | {experimenter} | {status_cn} | {tags} | {conclusion}"
            )

        exp_list_text = "\n".join(lines)
        system_prompt = (
            "你是实验记录搜索引擎。根据用户的自然语言描述，从实验列表中找出最匹配的实验。\n"
            "理解以下类型的查询：\n"
            "- 时间指代：'上周一'='最近一周'，'上个月'='30天前'，'最近'=按日期排序\n"
            "- 人员指代：'老张'='experimenter含张'，'我做的'=忽略\n"
            "- 状态指代：'失败的那个'='status=failed'，'成功的'='status=done且results有值'\n"
            "- 材料指代：'ZnO那个'='材料含ZnO'\n"
            "- 性能指代：'降解率最高的'='results中降解率数值最大的'\n\n"
            "严格返回 JSON 数组（不要包含在 markdown 代码块中）：\n"
            '[{"id": "2026-K7D2-001", "score": 0.95, "reason": "原因"}, ...]\n'
            "按匹配度降序排列，最多返回5个。score 0-1，0.3以下不要返回。\n"
            "如果没有匹配的实验，返回空数组 []。"
        )
        user_prompt = f"实验列表：\n{exp_list_text}\n\n用户查询：{query}\n\n请返回最匹配的实验 ID 列表（JSON 数组）："

        raw = loop.llm.analyze(system_prompt=system_prompt, user_prompt=user_prompt, temperature=0.1)

        # 容错解析
        try:
            results = json.loads(raw.strip())
            if isinstance(results, list):
                return results[:5]
        except json.JSONDecodeError:
            m = re.search(r'\[[\s\S]*\]', raw)
            if m:
                try:
                    results = json.loads(m.group(0))
                    if isinstance(results, list):
                        return results[:5]
                except json.JSONDecodeError:
                    pass
        return []

    def _fuzzy_search(self, query: str, loop: "AgentLoop", include_archived: bool = False) -> list[dict]:
        """本地关键词搜索（含实验 ID）"""
        if not query or len(query) < 2:
            return []
        all_exps = loop.store.list_all_full(include_archived=include_archived)
        results = []
        text_lower = query.lower()
        has_cjk = any('一' <= c <= '鿿' for c in query)

        for exp in all_exps:
            score = 0.0
            exp_id = (exp.get("id") or "").lower()
            title = (exp.get("title") or "").lower()
            tags = " ".join(exp.get("tags") or []).lower()
            purpose = (exp.get("purpose") or "")[:200].lower()
            mat_names = " ".join(
                m.get("name", "") for m in (exp.get("materials") or [])
                if isinstance(m, dict)
            ).lower()
            searchable = f"{exp_id} {title} {tags} {purpose} {mat_names}"

            if has_cjk:
                tokens = [text_lower]
                for i in range(len(text_lower) - 1):
                    tokens.append(text_lower[i:i + 2])
            else:
                tokens = text_lower.split()

            for token in tokens:
                if len(token) >= 2 and token in searchable:
                    score += 0.25

            for tag in (exp.get("tags") or []):
                if tag.lower() in text_lower:
                    score += 0.3

            if score >= 0.2:
                results.append({
                    "id": exp.get("id"),
                    "title": exp.get("title", ""),
                    "date": exp.get("date", ""),
                    "tags": exp.get("tags", []),
                    "score": min(score, 0.99),
                })

        results.sort(key=lambda r: r["score"], reverse=True)
        return results[:5]

    # -- Step 1.3: update_schema --

    def _update_schema(self, args: dict, loop: "AgentLoop") -> dict:
        """纯写入：合并 fields → 生成 Schema 状态 → 注入 messages"""
        if loop._schema_context is None:
            return {"error": "not_in_record_mode",
                    "message": "update_schema 只在记录实验时可用。"}
        fields = args.get("fields", {})

        # 追踪 modified_values：首次触及的字段记录旧值
        for key in fields:
            if key not in loop.modified_values:
                old_val = loop._schema_context.get(key)
                if isinstance(old_val, (list, dict)):
                    loop.modified_values[key] = deepcopy(old_val)
                else:
                    loop.modified_values[key] = old_val

        merge_context(loop._schema_context, fields)

        # 推断 experiment_type（从 tags 中）
        if loop.experiment_type == "other":
            tags = loop._schema_context.get("tags", [])
            for tag in tags:
                if tag in ("photocatalysis", "hydrothermal", "sol-gel",
                           "spin-coating", "ball-milling",
                           "electrochemistry", "xrd", "perovskite-solar"):
                    loop.experiment_type = tag
                    break

        # 如果当前在 analyze 线程中，先结束它再开始 record
        if loop.thread.id:
            for m in loop.history:
                if f"thread_begin id={loop.thread.id} type=analyze" in (m.get("content") or ""):
                    loop._append_history({"role": "system",
                        "content": f"[系统内部] thread_end id={loop.thread.id}"})
                    loop.thread_store.set_active_thread(None)
                    loop.thread.id = None
                    loop.thread.type = None
                    break

        # 生成 Schema 状态并注入 messages
        status_msg = loop._build_schema_status()
        loop._append_history({
            "role": "system",
            "content": status_msg,
        })

        return {
            "status": "ok",
            "updated_fields": list(fields.keys()),
        }


# ============================================================================
# Step 1.8 / 1.9 / 1.10: AgentLoop
# ============================================================================

class AgentLoop:
    """基于 tool calling 的对话循环"""

    def __init__(self, llm_client, experiment_store, *,
                 tool_executor: "ToolExecutor | None" = None,
                 thread_store=None, update_log_store=None,
                 favorites_store=None, analysis_store=None,
                  analysis_svc=None, extraction_svc=None,
                  attachment_store=None,
                  context_compression_trigger_tokens: int = _DEFAULT_CONTEXT_COMPRESSION_TRIGGER_TOKENS,
                  context_compression_chunk_tokens: int = _DEFAULT_CONTEXT_COMPRESSION_CHUNK_TOKENS):
        self.llm = llm_client
        self.store = experiment_store
        self._schema_context = None  # 16-field dict — only non-None in record mode
        self.history = []           # [{role, content, tool_calls?, tool_call_id?}]
        self.references = []        # 已加载的引用实验 ID
        self.experiment_type = "other"
        self.turn_count = 0
        if tool_executor is not None:
            self.tools = tool_executor
        else:
            self.tools = ToolExecutor(experiment_store, update_log_store=update_log_store,
                                      favorites_store=favorites_store,
                                      analysis_store=analysis_store,
                                      attachment_store=attachment_store)
        self._generated_preview = None   # generate_record 工具产出
        self._generated_notes = None
        self._llm_call_seq = 0      # LLM 调用全局序号（跨 turn 递增）

        # 会话ID：全新启动时生成，重启时从 _current_state.yaml 恢复
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 已写入压缩归档表的原始消息数量；恢复后继续编号，避免重复归档。
        self._compressed_history_count = 0
        # 压缩摘要只属于当前主聊天会话，绝不能写入账号级上下文或被子 Agent 继承。
        self._session_summary = ""
        self._compressed_until_sequence = -1
        self._context_compression_trigger_tokens = max(10_000, int(context_compression_trigger_tokens))
        requested_chunk = max(1_000, int(context_compression_chunk_tokens))
        self._context_compression_chunk_tokens = min(
            requested_chunk, self._context_compression_trigger_tokens - 1_000,
        )

        # 冷存储：被压缩裁掉的消息写入 _history/{user_id}/{session_id}.jsonl
        _uid = thread_store._uid() if thread_store else ""
        _cold_dir = Path(experiment_store.path).parent / "_history" / (_uid or "_anonymous")
        _cold_dir.mkdir(parents=True, exist_ok=True)
        self._cold_store_path = _cold_dir / f"{self.session_id}.jsonl"

        # 线程系统 + 子Agent 标记 + 服务引用
        self.thread_store = thread_store
        self.update_log_store = update_log_store
        self.thread = ThreadState()
        self.child = ChildContext()
        self.modified_values = {}
        self.analysis_svc = analysis_svc
        self.extraction_svc = extraction_svc
        self.attachment_store = attachment_store
        self._data_freshness_context = ""
        self._tool_batch_active = False
        self._deferred_history: list[dict[str, Any]] = []
        self._deferred_thread_ends: list[tuple[str, str]] = []

    @staticmethod
    def _message_time() -> str:
        return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")

    def _append_history(self, message: dict[str, Any]) -> dict[str, Any]:
        """统一记录新消息的服务端生成时间。"""
        entry = dict(message)
        entry.setdefault("created_at", self._message_time())
        # OpenAI/DeepSeek 要求 tool_calls 后只能紧跟对应 tool 消息。工具执行
        # 过程中产生的 system 状态改为本批 tool 结果写完后再追加。
        if self._tool_batch_active and entry.get("role") == "system":
            self._deferred_history.append(entry)
            return entry
        self.history.append(entry)
        return entry

    def _insert_history(self, index: int, message: dict[str, Any]) -> dict[str, Any]:
        entry = dict(message)
        entry.setdefault("created_at", self._message_time())
        self.history.insert(index, entry)
        return entry

    def _flush_deferred_history(self) -> None:
        """在一整批 tool 结果之后写入工具产生的内部状态。"""
        deferred, self._deferred_history = self._deferred_history, []
        self.history.extend(deferred)
        ends, self._deferred_thread_ends = self._deferred_thread_ends, []
        for thread_id, produced_id in ends:
            if self.thread.id == thread_id:
                self._finish_thread_end(thread_id, produced_id)

    def _repair_tool_history(self) -> bool:
        """修复旧状态中被 system 消息或中断调用打断的 tool-call 批次。"""
        source = [dict(message) for message in self.history if isinstance(message, dict)]
        repaired: list[dict[str, Any]] = []
        changed = len(source) != len(self.history)
        index = 0
        while index < len(source):
            message = source[index]
            calls = message.get("tool_calls") if message.get("role") == "assistant" else None
            call_ids = [call.get("id") for call in calls or [] if isinstance(call, dict) and call.get("id")]
            if not call_ids:
                repaired.append(message)
                index += 1
                continue

            expected = set(call_ids)
            found: dict[str, dict[str, Any]] = {}
            delayed: list[dict[str, Any]] = []
            cursor = index + 1
            while cursor < len(source):
                candidate = source[cursor]
                role = candidate.get("role")
                if role in ("assistant", "user"):
                    break
                if role == "tool" and candidate.get("tool_call_id") in expected and candidate.get("tool_call_id") not in found:
                    found[candidate["tool_call_id"]] = candidate
                elif role == "tool":
                    # 孤立或重复的 tool 消息不能发送给 OpenAI。保留原始内容以便
                    # 模型了解旧状态，但把它降级为普通内部日志。
                    delayed.append({
                        "role": "system",
                        "content": "[系统内部] 已归档一条无匹配工具调用的旧工具结果："
                                   + str(candidate.get("content") or ""),
                        "created_at": candidate.get("created_at") or self._message_time(),
                    })
                else:
                    delayed.append(candidate)
                cursor += 1

            normalized_batch: list[dict[str, Any]] = []
            for call_id in call_ids:
                tool_message = found.get(call_id)
                if tool_message is None:
                    tool_message = {
                        "role": "tool", "tool_call_id": call_id,
                        "content": json.dumps({
                            "error": "interrupted_tool_call",
                            "message": "此前工具调用未完成，已在恢复会话时终止。",
                        }, ensure_ascii=False),
                        "created_at": self._message_time(),
                    }
                normalized_batch.append(tool_message)
            normalized_batch.extend(delayed)
            if source[index + 1:cursor] != normalized_batch:
                changed = True
            repaired.append(message)
            repaired.extend(normalized_batch)
            index = cursor

        # 处理没有归属到任何 assistant tool_calls 的遗留 tool 消息。
        # 此类消息同样会被 API 拒绝，不能原样保留在请求历史中。
        sanitized: list[dict[str, Any]] = []
        pending_ids: set[str] = set()
        for message in repaired:
            if message.get("role") == "assistant":
                calls = message.get("tool_calls") or []
                pending_ids = {
                    call.get("id") for call in calls
                    if isinstance(call, dict) and call.get("id")
                }
                sanitized.append(message)
            elif message.get("role") == "tool":
                call_id = message.get("tool_call_id")
                if call_id in pending_ids:
                    pending_ids.remove(call_id)
                    sanitized.append(message)
                else:
                    changed = True
                    sanitized.append({
                        "role": "system",
                        "content": "[系统内部] 已归档一条无匹配工具调用的旧工具结果："
                                   + str(message.get("content") or ""),
                        "created_at": message.get("created_at") or self._message_time(),
                    })
            else:
                sanitized.append(message)

        if changed:
            self.history = sanitized
        return changed

    def _persist_repaired_state(self) -> None:
        """持久化协议修复，不触发摘要压缩或额外模型调用。"""
        self._checkpoint_runtime_state()

    def _checkpoint_runtime_state(self) -> None:
        """保存完整运行态快照；用于工具批次结束等可恢复边界。"""
        if not self.thread_store:
            return
        try:
            if self.child.is_child:
                key = self.thread.id or self.child.exp_id
                if key:
                    self.thread_store.save_child_state(key, self.state_to_dict())
            else:
                self.thread_store.save_current_state(self.state_to_dict())
        except Exception:
            pass

    def _repair_history_before_llm(self) -> None:
        if self._repair_tool_history():
            self._persist_repaired_state()

    @staticmethod
    def _llm_history(history: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """移除展示/归档元数据，保留符合 OpenAI 协议的消息正文。"""
        result = []
        for message in history:
            item = {key: value for key, value in message.items()
                    if key not in ("created_at", "attachments")}
            attachments = message.get("attachments") or []
            if attachments and item.get("role") == "user":
                manifest = "\n".join(
                    f"- {a.get('name', '附件')} | {a.get('mime', 'unknown')} | sha256={a.get('sha256', '')}"
                    for a in attachments if isinstance(a, dict)
                )
                item["content"] = (item.get("content") or "") + "\n\n[用户上传的附件]\n" + manifest
            result.append(item)
        return result

    @staticmethod
    def _time_context(user_message: str) -> str:
        """仅在用户问题涉及时间时提供当前时间，避免普通回复复述时间戳。"""
        text = (user_message or "").lower()
        keywords = (
            "今天", "昨天", "明天", "后天", "本周", "上周", "下周", "现在", "时间", "日期",
            "几点", "何时", "多久", "几天", "几周", "几月", "去年", "今年", "明年",
            "today", "yesterday", "tomorrow", "time", "date", "when", "week", "month",
        )
        if not any(keyword in text for keyword in keywords):
            return ""
        now = datetime.now(timezone.utc).astimezone(ZoneInfo("Asia/Shanghai"))
        return (
            "[当前时间上下文] " + now.strftime("%Y-%m-%d %H:%M（Asia/Shanghai）")
            + "。仅在回答时间、日期或相对时间的问题时使用；普通回复不要主动提及此信息。"
        )

    def set_data_freshness_context(self, content: str) -> None:
        """设置本顶层用户回合的数据同步状态；只作为请求尾部的短暂系统上下文。"""
        self._data_freshness_context = str(content or "").strip()

    # -- 模式管理 --

    @property
    def mode(self) -> str:
        """当前对话模式: 'general' | 'record' | 'analyze'"""
        if not self.thread.id:
            return "general"
        return self.thread.type or "general"

    def _enter_record_mode(self) -> None:
        """初始化 Schema 上下文（仅 record 模式）。"""
        self._schema_context = deepcopy(DEFAULT_CONTEXT)

    def _exit_record_mode(self) -> None:
        """清理 Schema 上下文。"""
        self._schema_context = None

    def _get_active_tools(self) -> list[dict]:
        """返回当前模式或子 Agent 角色可用的工具列表。"""
        # 子 Agent 角色优先 —— 返回固定工具清单，不走 mode 推断
        if self.child.agent_role == "analysis_reviewer":
            return [
            TOOL_SEARCH_EXPERIMENTS, TOOL_READ_EXPERIMENT, TOOL_LIST_EXPERIMENTS,
                TOOL_READ_UPDATE_LOG, TOOL_SEARCH_CHAT_HISTORY, TOOL_READ_CHAT_HISTORY,
                TOOL_LIST_CHAT_SESSIONS, TOOL_BROWSE_CHAT_HISTORY,
                TOOL_READ_ANALYSIS, TOOL_END_THREAD,
            ]
        if self.child.agent_role == "exp_editor":
            return [
                TOOL_READ_EXPERIMENT, TOOL_READ_UPDATE_LOG, TOOL_MODIFY_EXPERIMENT,
                TOOL_MANAGE_ARCHIVE, TOOL_SEARCH_ATTACHMENTS, TOOL_READ_ATTACHMENT,
                TOOL_MANAGE_ATTACHMENT, TOOL_END_THREAD,
            ]

        # 父 Agent —— 按模式返回
        common = [
            TOOL_SEARCH_EXPERIMENTS, TOOL_READ_EXPERIMENT,
            TOOL_LIST_EXPERIMENTS, TOOL_MANAGE_CATEGORY, TOOL_READ_UPDATE_LOG,
            TOOL_SEARCH_CHAT_HISTORY, TOOL_READ_CHAT_HISTORY,
            TOOL_LIST_CHAT_SESSIONS, TOOL_BROWSE_CHAT_HISTORY, TOOL_END_THREAD,
            TOOL_SEARCH_ATTACHMENTS, TOOL_READ_ATTACHMENT, TOOL_MANAGE_ATTACHMENT,
            TOOL_MANAGE_MUSIC, TOOL_READ_ANALYSIS,
        ]
        if self.mode == "record":
            common.extend([TOOL_START_RECORD_THREAD, TOOL_UPDATE_SCHEMA,
                          TOOL_GENERATE_RECORD, TOOL_MODIFY_EXPERIMENT, TOOL_MANAGE_ARCHIVE])
        elif self.mode == "general":
            common.extend([TOOL_START_RECORD_THREAD, TOOL_START_ANALYZE_THREAD,
                          TOOL_MODIFY_EXPERIMENT, TOOL_MANAGE_ARCHIVE])
        elif self.mode == "analyze":
            # analyze 模式不包含 modify_experiment —— 分析者不应修改实验
            common.extend([TOOL_START_ANALYZE_THREAD, TOOL_SELECT_EXPERIMENTS,
                          TOOL_GENERATE_ANALYSIS])
        return common

    # -- 主循环 --

    def run(self, user_message: str = "", attachments: list[dict[str, Any]] | None = None,
            created_at: str | None = None) -> dict:
        """处理一条用户消息。返回 {type, message?, context}"""
        log = get_logger()
        if user_message:
            self.thread.current_turn_user_idx = len(self.history)
            entry = {"role": "user", "content": user_message}
            if created_at:
                entry["created_at"] = created_at
            if attachments:
                entry["attachments"] = attachments
            self._append_history(entry)
            self.turn_count += 1
            if log:
                agent = "child" if self.child.is_child else "parent"
                log.agent(agent, "user", user_message, exp=self.child.exp_id)

        consecutive_errors = 0
        last_tool = None
        _no_progress_count = 0  # Track rounds without update_schema/analyze

        while True:
            self._maybe_inject_thread_start()   # 循环顶部检查 flag
            self._repair_history_before_llm()

            # 构建 LLM 消息：稳定前缀在前，按请求变化的状态与时间置后，便于服务端提示词缓存。
            messages = [
                {"role": "system", "content": build_system_prompt(self.child.agent_role)},
            ]
            if self._session_summary:
                messages.append({"role": "system", "content": f"[当前会话历史摘要]\n{self._session_summary}"})
            messages.extend(self._llm_history(self.history))
            # 请求层：record 模式下追加实时 Schema 状态
            if self.mode == "record" and self._schema_context is not None:
                messages.append({"role": "system",
                                "content": self._build_schema_status()})
            # 请求层：追加线程状态（始终在末尾）
            messages.append({"role": "system",
                            "content": self._build_thread_status()})
            if self._data_freshness_context:
                messages.append({"role": "system", "content": self._data_freshness_context})
            time_context = self._time_context(user_message)
            if time_context:
                messages.append({"role": "system", "content": time_context})
            self._llm_call_seq += 1
            seq = self._llm_call_seq

            # ---- 日志: LLM 请求 ----

            try:
                response = self.llm.chat(
                    messages=messages,
                    tools=self._get_active_tools(),
                    temperature=0.3,
                    reasoning_effort=self._reasoning_effort(),
                )
            except Exception as e:
                # 回退 history 到最近一条 user 消息，清理其后的所有残留
                cut = len(self.history)
                for i in range(len(self.history) - 1, -1, -1):
                    if self.history[i].get("role") == "user":
                        cut = i + 1
                        break
                del self.history[cut:]
                self.turn_count = sum(1 for m in self.history if m.get("role") == "user")
                self._append_history({"role": "system",
                    "content": f"[系统内部] LLM 调用失败（已重试3次）: {str(e)[:200]}"})
                if log:
                    log.system("error", "llm_call_failed", error=str(e)[:200])
                    agent = "child" if self.child.is_child else "parent"
                    log.agent(agent, "assistant",
                        "抱歉，AI 服务暂时不可用（已自动重试3次）。请稍后重试。",
                        exp=self.child.exp_id)
                self._save_runtime_state()
                return {"type": "reply",
                        "message": "抱歉，AI 服务暂时不可用（已自动重试3次）。请稍后重试。",
                        "context": self._schema_context}

            resp_content = response.content
            resp_tool_calls = response.tool_calls
            _reasoning = response.reasoning

            # ---- 日志: LLM 响应 ----

            # 纯文本 → 不再调工具，直接返回
            if resp_content and not resp_tool_calls:
                entry = {"role": "assistant", "content": resp_content}
                if _reasoning:
                    entry["reasoning_content"] = _reasoning
                self._append_history(entry)
                if log:
                    agent = "child" if self.child.is_child else "parent"
                    log.agent(agent, "assistant", resp_content, exp=self.child.exp_id)
                self._maybe_inject_thread_start()   # return 前检查
                self._check_thread_cancellation(_no_progress_count)
                self._save_runtime_state()
                return {"type": "reply", "message": resp_content,
                        "context": self._schema_context}

            # 调用了工具
            # 记录 assistant 文本（工具调用前的说明文字，只记一次）
            if log and resp_content:
                ag = "child" if self.child.is_child else "parent"
                tc_names = [tc["function"]["name"] for tc in (resp_tool_calls or [])]
                log.agent(ag, "assistant", resp_content, tool_calls=tc_names, exp=self.child.exp_id)

            tool_items = []
            has_record_tool = False
            for tc in (resp_tool_calls or []):
                name = tc["function"]["name"]
                raw_args_str = tc["function"]["arguments"]
                if name in ("update_schema", "analyze"):
                    has_record_tool = True
                try:
                    args, parse_error = json.loads(raw_args_str), None
                except (TypeError, json.JSONDecodeError) as exc:
                    args, parse_error = None, str(exc)
                if parse_error is None and not isinstance(args, dict):
                    args, parse_error = None, "工具参数必须是 JSON 对象"
                tool_items.append((tc, name, raw_args_str, args, parse_error))

            # 同一次模型响应的多个 tool call 必须共享同一条 assistant 消息。
            # 逐个写入会重复 resp_content，也会破坏标准的 tool call 对应关系。
            entry = {
                "role": "assistant",
                "content": resp_content or None,
                "tool_calls": [{
                    "id": tc["id"],
                    "type": "function",
                    "function": {"name": name, "arguments": raw_args_str},
                } for tc, name, raw_args_str, _, _ in tool_items],
            }
            if _reasoning:
                entry["reasoning_content"] = _reasoning
            self._append_history(entry)

            pause_item = None
            too_many_errors = False
            analysis_timeout = None
            self._tool_batch_active = True
            try:
                for tc, name, raw_args_str, args, parse_error in tool_items:
                    if pause_item:
                        result = {"error": "skipped_due_to_pause",
                                  "message": "本轮已有暂停型工具调用，后续调用未执行。"}
                    elif parse_error:
                        result = {"error": "invalid_tool_arguments",
                                  "message": f"工具参数不是合法 JSON：{parse_error}"}
                    else:
                        try:
                            result = self.tools.execute(name, args, self)
                        except Exception as exc:
                            result = {"error": "execution_failed",
                                      "message": str(exc)[:300]}
                    if log:
                        ag = "child" if self.child.is_child else "parent"
                        log.tool(ag, name, "error" not in result,
                                 exp=self.child.exp_id, **_tool_log_summary(name, args or {}, result))
                    self._append_history({"role": "tool", "tool_call_id": tc["id"],
                                          "content": json.dumps(result, ensure_ascii=False)})
                    if "error" in result:
                        consecutive_errors = consecutive_errors + 1 if name == last_tool else 1
                        last_tool = name
                        too_many_errors = too_many_errors or consecutive_errors >= 3
                        if name == "generate_analysis" and result.get("error") == "analysis_timeout":
                            analysis_timeout = result
                    else:
                        consecutive_errors, last_tool = 0, None
                    if result.get("pause") and pause_item is None:
                        pause_item = (name, args or {}, result)
            finally:
                self._tool_batch_active = False
                self._flush_deferred_history()
                # 工具结果及其延迟的 thread_end 已完整写入 history；此时刷新也可恢复。
                self._checkpoint_runtime_state()

            if too_many_errors:
                self._append_history({"role": "assistant", "content": "抱歉，处理请求时遇到技术问题。请换个方式描述。"})
                self._maybe_inject_thread_start()
                self._save_runtime_state()
                return {"type": "reply", "message": "抱歉，处理请求时遇到技术问题。请换个方式描述。",
                        "context": self._schema_context}

            if analysis_timeout:
                message = analysis_timeout["message"] + "请稍后手动重新生成报告。"
                self._append_history({"role": "assistant", "content": message})
                self._save_runtime_state()
                return {"type": "reply", "message": message, "context": self._schema_context}

            if pause_item:
                name, args, result = pause_item
                self._maybe_inject_thread_start()
                if not has_record_tool and self.thread.id:
                    _no_progress_count += 1
                self._check_thread_cancellation(_no_progress_count)
                if name == "generate_record":
                    if self.thread.id and not self.child.is_child:
                        exp_id = self._generated_preview.get("id", "") if self._generated_preview else ""
                        self._maybe_inject_thread_end(exp_id)
                    if self._generated_preview is None:
                        return {"type": "reply", "message": "生成失败，请重试或补充更多信息。",
                                "context": self._schema_context}
                    self._save_runtime_state()
                    return {"type": result.get("response_type", "generate"), "message": "实验记录已生成并保存。",
                            "state": self.state_to_dict() if result.get("include_state") else None,
                            "preview": self._generated_preview, "notes": self._generated_notes,
                            "context": self._schema_context}
                self._save_runtime_state()
                return {"type": "reply", "message": result.get("message") or resp_content or "请在面板中选择实验。",
                        "context": self._schema_context}

            # 更新无进展计数
            if not has_record_tool and self.thread.id:
                _no_progress_count += 1
            else:
                _no_progress_count = 0

            # 其他工具执行完 → 继续循环

    # -- 流式主循环 --

    def run_stream(self, user_message: str = "", attachments: list[dict[str, Any]] | None = None,
                   created_at: str | None = None) -> Generator[dict[str, Any], None, None]:
        """和 run() 逻辑相同，但通过 Generator yield SSE 事件实现流式输出。"""
        from lib.llm import StreamEvent

        log = get_logger()
        if user_message:
            self.thread.current_turn_user_idx = len(self.history)
            entry = {"role": "user", "content": user_message}
            if created_at:
                entry["created_at"] = created_at
            if attachments:
                entry["attachments"] = attachments
            self._append_history(entry)
            self.turn_count += 1
            if log:
                agent = "child" if self.child.is_child else "parent"
                log.agent(agent, "user", user_message, exp=self.child.exp_id)

        consecutive_errors = 0
        last_tool = None
        _no_progress_count = 0

        while True:
            self._maybe_inject_thread_start()
            self._repair_history_before_llm()

            messages = [{"role": "system", "content": build_system_prompt(self.child.agent_role)}]
            if self._session_summary:
                messages.append({"role": "system", "content": f"[当前会话历史摘要]\n{self._session_summary}"})
            messages.extend(self._llm_history(self.history))
            if self.mode == "record" and self._schema_context is not None:
                messages.append({"role": "system", "content": self._build_schema_status()})
            messages.append({"role": "system", "content": self._build_thread_status()})
            if self._data_freshness_context:
                messages.append({"role": "system", "content": self._data_freshness_context})
            time_context = self._time_context(user_message)
            if time_context:
                messages.append({"role": "system", "content": time_context})
            self._llm_call_seq += 1
            seq = self._llm_call_seq

            try:
                stream = self.llm.chat_stream(
                    messages=messages,
                    tools=self._get_active_tools(),
                    temperature=0.3,
                    reasoning_effort=self._reasoning_effort(),
                )
            except Exception as e:
                cut = len(self.history)
                for i in range(len(self.history) - 1, -1, -1):
                    if self.history[i].get("role") == "user":
                        cut = i + 1
                        break
                del self.history[cut:]
                self.turn_count = sum(1 for m in self.history if m.get("role") == "user")
                self._append_history({"role": "system",
                    "content": f"[系统内部] LLM 调用失败: {str(e)[:200]}"})
                yield from self._save_runtime_state_stream()
                yield {"event": "error", "message": "LLM 调用失败，请稍后重试。"}
                return

            resp_content = ""
            resp_tool_calls = None
            _reasoning = ""
            current_tool = ""

            # 消费流事件
            try:
                while True:
                    try:
                        event = next(stream)
                    except StopIteration as exc:
                        resp = exc.value
                        resp_content = resp.content
                        resp_tool_calls = resp.tool_calls
                        _reasoning = resp.reasoning
                        break

                    if event.type == "text":
                        yield {"event": "text", "content": event.content}
                    elif event.type == "tool_call":
                        if event.tool_name and event.tool_name != current_tool:
                            current_tool = event.tool_name
                            yield {"event": "tool", "name": current_tool}
            except Exception as e:
                yield {"event": "error", "message": str(e)[:200]}
                return


            # 纯文本 → 返回
            if resp_content and not resp_tool_calls:
                entry = {"role": "assistant", "content": resp_content}
                if _reasoning:
                    entry["reasoning_content"] = _reasoning
                self._append_history(entry)
                self._check_thread_cancellation(_no_progress_count)
                yield from self._save_runtime_state_stream()
                yield {"event": "done", "type": "reply", "message": resp_content,
                       "context": self._schema_context}
                return

            # 调用了工具
            if log and resp_content:
                ag = "child" if self.child.is_child else "parent"
                tc_names = [tc["function"]["name"] for tc in (resp_tool_calls or [])]
                log.agent(ag, "assistant", resp_content, tool_calls=tc_names, exp=self.child.exp_id)

            tool_items = []
            has_record_tool = False
            for tc in (resp_tool_calls or []):
                name = tc["function"]["name"]
                raw_args_str = tc["function"]["arguments"]
                if name in ("update_schema", "analyze"):
                    has_record_tool = True
                try:
                    args, parse_error = json.loads(raw_args_str), None
                except (TypeError, json.JSONDecodeError) as exc:
                    args, parse_error = None, str(exc)
                if parse_error is None and not isinstance(args, dict):
                    args, parse_error = None, "工具参数必须是 JSON 对象"
                tool_items.append((tc, name, raw_args_str, args, parse_error))

            entry = {
                "role": "assistant",
                "content": resp_content or None,
                "tool_calls": [{
                    "id": tc["id"],
                    "type": "function",
                    "function": {"name": name, "arguments": raw_args_str},
                } for tc, name, raw_args_str, _, _ in tool_items],
            }
            if _reasoning:
                entry["reasoning_content"] = _reasoning
            self._append_history(entry)

            pause_item = None
            too_many_errors = False
            analysis_timeout = None
            self._tool_batch_active = True
            try:
                for tc, name, raw_args_str, args, parse_error in tool_items:
                    if pause_item:
                        result = {"error": "skipped_due_to_pause",
                                  "message": "本轮已有暂停型工具调用，后续调用未执行。"}
                    elif parse_error:
                        result = {"error": "invalid_tool_arguments",
                                  "message": f"工具参数不是合法 JSON：{parse_error}"}
                    else:
                        try:
                            result = self.tools.execute(name, args, self)
                        except Exception as exc:
                            result = {"error": "execution_failed",
                                      "message": str(exc)[:300]}
                    if log:
                        ag = "child" if self.child.is_child else "parent"
                        log.tool(ag, name, "error" not in result,
                                 exp=self.child.exp_id, **_tool_log_summary(name, args or {}, result))
                    tool_entry = self._append_history({
                        "role": "tool", "tool_call_id": tc["id"],
                        "content": json.dumps(result, ensure_ascii=False),
                    })
                    # 先让 history 完整，再通知浏览器工具已完成；刷新/断连不会留下半批记录。
                    yield {"event": "tool_done", "name": name,
                           "tool_call_id": tc["id"], "created_at": tool_entry.get("created_at"),
                           "args": args or {}, "result": result}
                    if "error" in result:
                        consecutive_errors = consecutive_errors + 1 if name == last_tool else 1
                        last_tool = name
                        too_many_errors = too_many_errors or consecutive_errors >= 3
                        if name == "generate_analysis" and result.get("error") == "analysis_timeout":
                            analysis_timeout = result
                    else:
                        consecutive_errors, last_tool = 0, None
                    if result.get("pause") and pause_item is None:
                        pause_item = (name, args or {}, result)
            finally:
                self._tool_batch_active = False
                self._flush_deferred_history()
                # 流式请求的下一段模型回复尚未结束前，也必须有可恢复快照。
                self._checkpoint_runtime_state()

            if too_many_errors:
                self._append_history({"role": "assistant", "content": "抱歉，处理请求时遇到技术问题。"})
                yield from self._save_runtime_state_stream()
                yield {"event": "done", "type": "reply", "message": "抱歉，处理请求时遇到技术问题。"}
                return

            if analysis_timeout:
                message = analysis_timeout["message"] + "请稍后手动重新生成报告。"
                self._append_history({"role": "assistant", "content": message})
                yield from self._save_runtime_state_stream()
                yield {"event": "done", "type": "reply", "message": message}
                return

            if pause_item:
                name, args, result = pause_item
                if name == "generate_record":
                    if self.thread.id and not self.child.is_child:
                        exp_id = self._generated_preview.get("id", "") if self._generated_preview else ""
                        self._maybe_inject_thread_end(exp_id)
                    if self._generated_preview is None:
                        yield from self._save_runtime_state_stream()
                        yield {"event": "done", "type": "reply", "message": "生成失败，请重试或补充更多信息。"}
                        return
                    yield from self._save_runtime_state_stream()
                    yield {"event": "done", "type": "generate", "message": "实验记录已生成并保存。",
                           "preview": self._generated_preview, "notes": self._generated_notes}
                    return
                yield from self._save_runtime_state_stream()
                yield {"event": "done", "type": "reply", "message": result.get("message") or resp_content or "请在面板中选择实验。"}
                return

            if not has_record_tool and self.thread.id:
                _no_progress_count += 1
            else:
                _no_progress_count = 0

    # -- Step 1.4: Schema 状态摘要 --

    def _build_schema_status(self) -> str:
        """生成 Schema 状态摘要，注入 messages。LLM 直接读这个判断缺什么。"""
        schema_fields = [
            ("title", "标题"), ("date", "日期"), ("experimenter", "实验者"),
            ("status", "状态"), ("tags", "标签"), ("purpose", "目的"),
            ("materials", "材料"), ("equipment", "设备"),
            ("experimental_plan", "方案"), ("sop", "步骤"),
            ("process_parameters", "参数"), ("observations", "观察"),
            ("characterization", "表征"), ("results", "结果"),
            ("conclusion", "结论"), ("next_steps", "下一步"),
        ]

        filled = []
        missing = []
        for key, label in schema_fields:
            val = self._schema_context.get(key) if self._schema_context else None
            if _is_filled(val):
                filled.append(f"{label}({_brief(val)})")
            else:
                missing.append(label)

        lines = [
            f"[Schema状态] 已填充 {len(filled)}/{len(schema_fields)} 字段",
            f"已填: {', '.join(filled) if filled else '(无)'}",
            f"缺失: {', '.join(missing) if missing else '(无)'}",
        ]
        if missing and len(filled) / len(schema_fields) >= 0.7:
            lines.append("提示: 缺失项多为补充字段，可考虑结束收集。")

        return "\n".join(lines)

    # -- 核心字段检查 --

    def _build_notes_from_context(self) -> str:
        """从 context 生成自然语言实验描述（Python 模板，不调 LLM）"""
        ctx = self._schema_context or {}
        parts = []
        if ctx.get("title"):
            parts.append(f"实验标题: {ctx['title']}")
        if ctx.get("date"):
            parts.append(f"日期: {ctx['date']}")
        if ctx.get("experimenter"):
            parts.append(f"实验者: {ctx['experimenter']}")
        tags = ctx.get("tags", [])
        if tags:
            parts.append(f"标签: {', '.join(str(t) for t in tags)}")
        status_val = ctx.get("status", "")
        if status_val and status_val != "planned":
            status_cn = {"planned": "计划中", "running": "进行中", "done": "已完成",
                         "failed": "失败", "repeated": "重复"}.get(status_val, status_val)
            parts.append(f"状态: {status_cn}")
        if ctx.get("purpose"):
            parts.append(f"实验目的: {ctx['purpose']}")
        materials = ctx.get("materials", [])
        if materials:
            lines = ["材料与试剂:"]
            for m in materials:
                if isinstance(m, dict):
                    name = m.get("name", "")
                    purity = f", 纯度 {m['purity']}" if m.get("purity") else ""
                    vendor = f", {m['vendor']}" if m.get("vendor") else ""
                    amount = f", {m['amount']}" if m.get("amount") else ""
                    lines.append(f"  - {name}{purity}{vendor}{amount}")
            parts.append("\n".join(lines))
        equipment = ctx.get("equipment", [])
        if equipment:
            lines = ["仪器设备:"]
            for e in equipment:
                if isinstance(e, dict):
                    lines.append(f"  - {e.get('device', '')}")
            parts.append("\n".join(lines))
        sop = ctx.get("sop", [])
        if sop:
            lines = ["实验步骤:"]
            for i, s in enumerate(sop, 1):
                lines.append(f"  {i}. {s}")
            parts.append("\n".join(lines))
        exp_plan = ctx.get("experimental_plan", [])
        if exp_plan:
            lines = ["实验方案:"]
            for i, p in enumerate(exp_plan, 1):
                if isinstance(p, dict):
                    group = p.get("group", "")
                    condition = p.get("condition", "")
                    expected = f", 预期{p.get('expected', '')}" if p.get("expected") else ""
                    lines.append(f"  {i}. 组'{group}': {condition}{expected}")
            parts.append("\n".join(lines))
        params = ctx.get("process_parameters", [])
        if params:
            lines = ["过程参数:"]
            for p in params:
                if isinstance(p, dict):
                    lines.append(f"  - {p.get('parameter', '')}: {p.get('setpoint', '')}")
            parts.append("\n".join(lines))
        chara = ctx.get("characterization", [])
        if chara:
            lines = ["表征手段:"]
            for c in chara:
                if isinstance(c, dict):
                    lines.append(f"  - {c.get('method', '')}")
            parts.append("\n".join(lines))
        results = ctx.get("results", {})
        if isinstance(results, dict):
            if results.get("qualitative"):
                parts.append(f"定性结果: {results['qualitative']}")
            kd = results.get("key_data", [])
            if kd:
                lines = ["关键数据:"]
                for k in kd:
                    if isinstance(k, dict):
                        lines.append(f"  - {k.get('metric', '')}: {k.get('value', '')}")
                parts.append("\n".join(lines))
        obs = ctx.get("observations", {})
        if isinstance(obs, dict):
            items = obs.get("items", [])
            if items:
                parts.append("异常观察: " + "; ".join(str(i) for i in items))
        if ctx.get("conclusion"):
            parts.append(f"结论: {ctx['conclusion']}")
        if ctx.get("next_steps"):
            nss = ctx["next_steps"]
            if isinstance(nss, list):
                parts.append("下一步: " + "; ".join(str(s) for s in nss))
        return "\n\n".join(parts) if parts else "（无实验描述）"

    def _core_fields_filled(self) -> bool:
        """检查核心字段是否已填充。"""
        CORE_BY_TYPE = {
            "photocatalysis": ["purpose", "materials", "process_parameters", "results"],
            "hydrothermal": ["purpose", "materials", "sop", "process_parameters", "results"],
            "sol-gel": ["purpose", "materials", "sop", "process_parameters", "results"],
            "spin-coating": ["purpose", "materials", "sop", "process_parameters", "results"],
            "ball-milling": ["purpose", "materials", "sop", "process_parameters", "results"],
            "electrochemistry": ["purpose", "materials", "process_parameters", "results"],
            "xrd": ["purpose", "materials", "process_parameters", "results"],
            "perovskite-solar": ["purpose", "materials", "sop", "process_parameters", "results"],
        }
        core = CORE_BY_TYPE.get(self.experiment_type,
                                ["purpose", "materials", "sop", "results"])
        return all(_is_filled((self._schema_context or {}).get(f)) for f in core)

    # -- 线程系统 --

    def _build_thread_guidance(self, thread_type: str) -> dict:
        """生成线程模式引导消息。"""
        if thread_type == "record":
            return {"role": "system",
                    "content": "你正在记录一条新实验。优先收集材料、步骤、参数、结果。追问缺失的关键字段。目标：generate_record。"}
        elif thread_type == "analyze":
            return {"role": "system",
                    "content": "你正在进行跨实验分析。先了解用户需求，用 search_experiments 或 list_experiments 缩小范围，再用 select_experiments 让用户勾选实验。选择确认后，以自然语言确认分析角度；不要为了生成报告而读取实验全文。需求明确后调用 generate_analysis，只传入用户确认的分析需求；系统会将已选实验的完整记录、附件和更新日志直接交给分析 Worker。"}
        return {"role": "system", "content": ""}

    def _build_thread_status(self) -> str:
        """生成当前线程状态声明。每轮 LLM 请求注入，不入 history。"""
        # 子 Agent 角色覆盖 —— 不依赖 _thread_type
        if self.child.agent_role == "analysis_reviewer":
            return (
                "[系统状态] 你正在审阅一份已完成的分析报告。"
                "可用工具：read_experiment（查看报告中引用的实验）、search_experiments、"
                "read_update_log、read_analysis（读取当前报告完整正文）。分析报告已归档，不可修改。"
                "不要使用 start_analyze_thread、select_experiments、generate_analysis——这些属于分析创建阶段。"
            )
        if self.child.agent_role == "exp_editor":
            return (
                f"[系统状态] 你正在修改已完成的实验 {self.child.exp_id}，只能处理这一条实验。"
                "修改前先用 read_experiment 读取磁盘最新数据和 revision（不要依赖对话记忆）。"
                "修改用 modify_experiment 工具直接执行，会自动保存和记录日志。"
                "附件只能读取当前实验已关联的文件或本次对话上传的文件，并且只能关联到当前实验。"
            )

        if not self.thread.id:
            return (
                "[系统状态] 自由模式。"
                "你可回答查询、管理收藏、闲聊。"
                "用户要记录新实验时调用 start_record_thread，"
                "要跨实验分析时调用 start_analyze_thread。"
            )
        if self.thread.type == "record":
            return (
                "[系统状态] record 线程进行中。"
                "持续收集实验信息，缺失关键字段时追问。目标：generate_record。"
            )
        if self.thread.type == "analyze":
            return (
                "[系统状态] analyze 线程进行中。"
                "深入讨论，使用 search_experiments + read_experiment(as_reference=true) + 自身推理。"
                "目标：输出分析报告。"
            )
        return "[系统状态] 自由模式。"

    def _maybe_inject_thread_start(self) -> None:
        """analyze 工具触发时注入线程标记。record 线程由 start_record_thread 工具直接处理。"""
        if not self.thread.pending_start or not self.thread_store:
            return
        thread_type = self.thread.pending_start
        self.thread.pending_start = None
        thread_id = self.thread_store.next_id()
        self.thread.id = thread_id
        self.thread.type = thread_type
        self.thread_store.set_active_thread(thread_id)
        if thread_type == "record":
            self._enter_record_mode()
        begin = {"role": "system", "content": f"[系统内部] thread_begin id={thread_id} type={thread_type}"}
        insert_pos = self.thread.current_turn_user_idx + 1
        self._insert_history(insert_pos, begin)
        guidance = self._build_thread_guidance(thread_type)
        if guidance.get("content"):
            self._insert_history(insert_pos + 1, guidance)
        self.thread_store.create(thread_type, [begin, guidance] if guidance.get("content") else [begin])
        log = get_logger()
        if log:
            log.operation("thread_start", agent="parent", thread=thread_id, type=thread_type)

    def _maybe_inject_thread_end(self, produced_id: str) -> None:
        """注入线程结束标记 + 提取 messages → 写线程文件 + 更新索引 + 重置上下文。"""
        if not self.thread.id or not self.thread_store:
            return
        thread_id = self.thread.id
        end = {"role": "system",
               "content": f"[系统内部] thread_end id={thread_id} product={produced_id}"}
        self._append_history(end)
        if self._tool_batch_active:
            self._deferred_thread_ends.append((thread_id, produced_id))
            return
        self._finish_thread_end(thread_id, produced_id)

    def _finish_thread_end(self, thread_id: str, produced_id: str) -> None:
        """结束标记已进入 history 后，再归档并重置线程状态。"""
        if self.thread.id != thread_id or not self.thread_store:
            return
        self._extract_and_save_thread(produced_id)
        self.thread_store.set_active_thread(None)
        # 统一日志
        log = get_logger()
        if log:
            agent = "child" if self.child.is_child else "parent"
            log.operation("thread_end", agent=agent, thread=thread_id, produced=produced_id)
        # 记录刚结束的线程ID，压缩时跳过它
        self.thread.last_ended_id = thread_id
        self.thread.id = None
        self.thread.type = None
        # 清理 Schema 状态和引用
        self._exit_record_mode()
        self.references = []
        self.experiment_type = "other"
        self.modified_values = {}

    def _extract_and_save_thread(self, produced_id: str) -> None:
        """提取 begin-end 标记间的 messages → 写入线程文件 + 更新索引。"""
        tid = self.thread.id
        # 找到 begin 标记位置
        begin_idx = None
        end_idx = None
        for i, m in enumerate(self.history):
            content = m.get("content") or ""
            if f"thread_begin id={tid}" in content:
                begin_idx = i
            elif f"thread_end id={tid}" in content:
                end_idx = i
                break
        if begin_idx is None or end_idx is None:
            return
        # 提取区间 messages（从触发用户消息开始，它位于 begin 标记之前）
        thread_msgs = self.history[begin_idx - 1:end_idx + 1]
        # 更新线程文件
        thread = self.thread_store.load(tid)
        if thread:
            thread["messages"] = thread_msgs
            thread["status"] = "done"
            thread["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if is_experiment_id(produced_id) and not thread.get("exp_generated"):
                thread["exp_generated"] = produced_id
                # 从首个 user 消息截取标题
                for m in thread_msgs:
                    if m.get("role") == "user":
                        first_user = m.get("content") or ""[:30]
                        thread["title"] = first_user
                        break
                thread["summary"] = f"生成{produced_id}"
            elif produced_id.startswith("ANAL-"):
                thread["anal_generated"] = produced_id
            # Title: if >= 3 turns, generate with LLM later (simplified for now)
            self.thread_store.save(thread)
            self.thread_store.update_index(thread)
    def _check_thread_cancellation(self, consecutive_no_progress: int) -> None:
        """检测线程是否需要取消。返回更新后的 consecutive_no_progress。"""
        if not self.thread.id:
            return
        # 简单实现：如果连续 3 轮无进展，自动取消
        # 注意：调用方负责维护 consecutive_no_progress 计数
        if consecutive_no_progress >= 3:
            tid = self.thread.id
            self._append_history({"role": "system",
                "content": f"[系统内部] thread_cancelled id={tid}"})
            # 移除 begin 标记
            for i, m in enumerate(self.history):
                if f"thread_begin id={tid}" in (m.get("content") or ""):
                    self.history.pop(i)
                    # 同时移除紧跟的引导消息
                    if i < len(self.history) and self.history[i].get("role") == "system":
                        content = self.history[i].get("content") or ""
                        if "正在记录" in content or "正在进行" in content:
                            self.history.pop(i)
                    break
            self.thread.id = None
            self.thread.type = None
            self._exit_record_mode()
            self.modified_values = {}
            log = get_logger()
            if log:
                agent = "child" if self.child.is_child else "parent"
                log.operation("thread_cancelled", agent=agent, thread=tid)

    # -- 子 Agent --

    @classmethod
    def create_child_agent(cls, parent_loop: "AgentLoop", thread_id: str) -> "AgentLoop":
        """从父 Agent 创建子 Agent，用于续接历史线程（修改已完成的实验）。"""
        thread = parent_loop.thread_store.load(thread_id)
        if not thread:
            raise ValueError(f"Thread {thread_id} not found")

        child = cls(
            parent_loop.llm,
            parent_loop.store,
            thread_store=parent_loop.thread_store,
            update_log_store=parent_loop.update_log_store,
            favorites_store=getattr(parent_loop.tools, 'favorites_store', None),
            analysis_store=getattr(parent_loop.tools, 'analysis_store', None),
        )
        # 子 Agent 只继承该线程的完整 messages（LLM 参考用）。
        for m in thread.get("messages", []):
            if m.get("role") != "system" or "[全局上下文]" not in (m.get("content") or ""):
                child.history.append(dict(m))
        # 记录初始 history 长度——前端只渲染此索引之后的消息
        child.child.initial_history_len = len(child.history)
        child.thread.id = thread_id
        child.child.is_child = True
        child.child.agent_role = "exp_editor"
        return child

    @classmethod
    def create_legacy_child_agent(cls, llm_client, store, exp_data: dict,
                                   thread_store=None, update_log_store=None,
                                   favorites_store=None,
                                   analysis_store=None, attachment_store=None) -> "AgentLoop":
        """为无线程关联的旧实验创建子 Agent，并注入 EXP 结构化数据。"""
        child = cls(llm_client, store,
                    thread_store=thread_store,
                    update_log_store=update_log_store,
                    favorites_store=favorites_store,
                    analysis_store=analysis_store,
                    attachment_store=attachment_store)
        # 注入 EXP 数据作为上下文
        child._append_history({
            "role": "system",
            "content": f"[当前实验数据]\n{json.dumps(exp_data, ensure_ascii=False, indent=2)}"
        })
        child.child.is_child = True
        child.child.is_legacy = True
        child.child.agent_role = "exp_editor"
        return child

    # -- 持久化: 每轮结束时实时保存 --



    def _save_runtime_state(self) -> None:
        """保存 AgentLoop 运行时状态。父 Agent 写 _current_state.yaml；子 Agent 写 child_state.yaml（不碰 _current_state.yaml）。"""
        if not self.thread_store:
            return
        try:
            # 先压缩再持久化；否则数据库会一直保留压缩前的大 history，
            # 页面刷新和下一轮恢复都会重新加载它。
            if not self.child.is_child:
                self._maybe_summarize()
            if self.child.is_child:
                # 子 Agent: 写独立 child_state.yaml，绝不覆盖父 Agent 的 _current_state.yaml
                key = self.thread.id or self.child.exp_id
                if key:
                    self.thread_store.save_child_state(key, self.state_to_dict())
            else:
                # 父 Agent: 写 _current_state.yaml
                self.thread_store.save_current_state(self.state_to_dict())
        except Exception:
            pass

    def _will_compress_current_session(self) -> bool:
        """只用于在流式连接中提前告知浏览器，实际判断仍由 _maybe_summarize 执行。"""
        return bool(
            self.thread_store and not self.child.is_child
            and self._estimate_request_tokens() >= self._context_compression_trigger_tokens
        )

    def _save_runtime_state_stream(self):
        """流式路径在压缩请求开始前先发出状态事件。"""
        if self._will_compress_current_session():
            yield {"event": "compression", "stage": "start"}
        self._save_runtime_state()

    # -- 上下文窗口管理 --

    def _reasoning_effort(self) -> str:
        value = str(getattr(self.llm, "default_reasoning_effort", "max") or "max").lower()
        return value if value in {"low", "medium", "high", "max"} else "max"

    @staticmethod
    def _estimate_tokens(text: str) -> int:
        """估算 token 数。中文/CJK ~1 char/token，英文 ~4 char/token。"""
        ideograph = 0
        ascii_chars = 0
        for c in text:
            cp = ord(c)
            if cp < 128:
                ascii_chars += 1
            elif (0x4E00 <= cp <= 0x9FFF       # CJK 基本区
                  or 0x3400 <= cp <= 0x4DBF     # CJK 扩展 A
                  or 0x20000 <= cp <= 0x2A6DF   # CJK 扩展 B
                  or 0xF900 <= cp <= 0xFAFF     # CJK 兼容汉字
                  or 0x3000 <= cp <= 0x303F     # CJK 标点
                  or 0xFF00 <= cp <= 0xFFEF     # 全角标点
                  or 0xAC00 <= cp <= 0xD7AF     # 韩文
                  or 0x3040 <= cp <= 0x309F     # 日文平假名
                  or 0x30A0 <= cp <= 0x30FF     # 日文片假名
            ):
                ideograph += 1
        other = len(text) - ideograph - ascii_chars
        return int(ideograph * 1.2 + ascii_chars * 0.25 + other * 0.8)

    @classmethod
    def _estimate_message_tokens(cls, message: dict[str, Any]) -> int:
        """按实际传给模型的完整消息对象估算，不能只看 content。"""
        try:
            text = json.dumps(message, ensure_ascii=False, separators=(",", ":"))
        except (TypeError, ValueError):
            text = str(message)
        return cls._estimate_tokens(text)

    def _estimate_request_tokens(self) -> int:
        """估算下一次 LLM 请求的全部组成，而非仅 history 正文。"""
        messages = [{"role": "system", "content": build_system_prompt(self.child.agent_role)}]
        if self._session_summary:
            messages.append({"role": "system", "content": f"[当前会话历史摘要]\n{self._session_summary}"})
        messages.extend(self._llm_history(self.history))
        if self.mode == "record" and self._schema_context is not None:
            messages.append({"role": "system", "content": self._build_schema_status()})
        messages.append({"role": "system", "content": self._build_thread_status()})
        if self._data_freshness_context:
            messages.append({"role": "system", "content": self._data_freshness_context})
        return sum(self._estimate_message_tokens(message) for message in messages)

    def _context_compression_keep_tokens(self) -> int:
        """按“触发阈值 - 每次整理量”保留最近完整对话。"""
        return self._context_compression_trigger_tokens - self._context_compression_chunk_tokens

    @staticmethod
    def _history_groups(history: list[dict[str, Any]]) -> list[tuple[int, int]]:
        """将 assistant tool_calls 与紧随的 tool result 保持为不可拆分的一组。"""
        groups = []
        i = 0
        while i < len(history):
            message = history[i]
            if message.get("role") == "assistant" and message.get("tool_calls"):
                call_ids = {call.get("id") for call in message["tool_calls"] if call.get("id")}
                j = i + 1
                while (j < len(history) and history[j].get("role") == "tool"
                       and history[j].get("tool_call_id") in call_ids):
                    j += 1
                groups.append((i, j))
                i = j
                continue
            groups.append((i, i + 1))
            i += 1
        return groups

    def _maybe_summarize(self) -> None:
        """仅压缩主 Agent 当前会话的旧消息；原文始终归档保留。"""
        if not self.thread_store or self.child.is_child:
            return
        new_msgs = self.history
        if self._estimate_request_tokens() < self._context_compression_trigger_tokens:
            return
        # 保留“触发阈值 - 整理量”的最近完整消息；tool call 与 result 必须一起保留。
        keep_start = len(new_msgs)
        kept_tokens = 0
        keep_limit = self._context_compression_keep_tokens()
        for start, end in reversed(self._history_groups(new_msgs)):
            group_tokens = sum(self._estimate_message_tokens(m) for m in new_msgs[start:end])
            # 单个工具调用组可能远大于保留额度。若把它也留下，会导致压缩后仍
            # 每轮超阈值；此时保留更新的完整回复，把这个已完成的调用组纳入摘要。
            if kept_tokens and kept_tokens + group_tokens > keep_limit:
                break
            kept_tokens += group_tokens
            keep_start = start
            if kept_tokens >= keep_limit:
                break
        to_summarize = new_msgs[:keep_start]
        if not to_summarize:
            return
        # 尝试 LLM 压缩；失败则保留完整 history，下次再试
        import json as _json
        try:
            # 传入全部待压缩消息，不截取尾部、条数或正文；请求过大时由调用失败
            # 路径保留原文，绝不以不完整摘要替代。
            text = _json.dumps(to_summarize, ensure_ascii=False, separators=(",", ":"))
            prior_summary = self._session_summary.strip()
            raw = self.llm.analyze(
                system_prompt="你是对话摘要助手。把已有摘要与以下完整的新压缩段合并为一份当前会话摘要。保留实验记录、修改操作、关键决策、工具调用及其结论。不要引用外部信息。用中文，控制在 500-2000 字。",
                user_prompt=(f"已有摘要：\n{prior_summary or '（无）'}\n\n"
                             f"请合并的完整对话：\n\n{text}"),
                temperature=0.2,
                request_timeout=_CONTEXT_COMPRESSION_REQUEST_TIMEOUT_SECONDS,
            )
            new_summary = raw[:2000]
        except Exception:
            return  # 压缩失败：history 完好，不裁剪，不写冷存储，下次再试
        # 压缩成功：先把完整原文写入 SQLite，成功后才能裁剪运行态。
        # JSONL 继续保留作兼容备份，但后续查找应以数据库为准。
        try:
            if hasattr(self.thread_store, "archive_compressed_history"):
                self.thread_store.archive_compressed_history(
                    self.session_id, self._compressed_history_count, to_summarize,
                )
            with open(self._cold_store_path, "a", encoding="utf-8") as f:
                for m in to_summarize:
                    f.write(_json.dumps(m, ensure_ascii=False) + "\n")
        except Exception:
            return
        self._compressed_history_count += len(to_summarize)
        self._compressed_until_sequence = self._compressed_history_count - 1
        self.history = new_msgs[keep_start:]
        self._session_summary = new_summary

    # -- 状态序列化 --

    def state_to_dict(self) -> dict:
        return {
            "context": self._schema_context,
            "references": self.references,
            "experiment_type": self.experiment_type,
            "turn_count": self.turn_count,
            "llm_call_seq": self._llm_call_seq,
            "history": [
                {k: v for k, v in m.items() if v is not None}
                for m in self.history
            ],
            "thread_id": self.thread.id,
            "_thread_type": self.thread.type,
            "_pending_thread_start": self.thread.pending_start,
            "_current_turn_user_idx": self.thread.current_turn_user_idx,
            "modified_values": dict(self.modified_values),
            "_session_id": self.session_id,
            "_compressed_history_count": self._compressed_history_count,
            "_session_summary": self._session_summary,
            "_compressed_until_sequence": self._compressed_until_sequence,
            "_is_child_agent": self.child.is_child,
            "_is_legacy": self.child.is_legacy,
            "_child_exp_id": self.child.exp_id,
            "_child_initial_history_len": self.child.initial_history_len,
            "_child_agent_role": self.child.agent_role,
        }

    @classmethod
    def from_dict(cls, llm_client, store, data: dict,
                  thread_store=None, update_log_store=None,
                  favorites_store=None, analysis_store=None,
                  analysis_svc=None, extraction_svc=None, attachment_store=None,
                  context_compression_trigger_tokens: int = _DEFAULT_CONTEXT_COMPRESSION_TRIGGER_TOKENS,
                  context_compression_chunk_tokens: int = _DEFAULT_CONTEXT_COMPRESSION_CHUNK_TOKENS) -> "AgentLoop":
        loop = cls(llm_client, store, thread_store=thread_store, update_log_store=update_log_store,
                   favorites_store=favorites_store, analysis_store=analysis_store,
                   analysis_svc=analysis_svc, extraction_svc=extraction_svc,
                   attachment_store=attachment_store,
                   context_compression_trigger_tokens=context_compression_trigger_tokens,
                   context_compression_chunk_tokens=context_compression_chunk_tokens)
        # 向后兼容：旧的 context 可能为空 dict（不是 None），按 None 处理
        ctx = data.get("context")
        if ctx and any(_is_filled(v) for v in ctx.values()):
            loop._schema_context = ctx
        else:
            loop._schema_context = None
        loop.references = data.get("references", [])
        loop.experiment_type = data.get("experiment_type", "other")
        loop.turn_count = data.get("turn_count", 0)
        loop._llm_call_seq = data.get("llm_call_seq", 0)
        history = data.get("history", [])
        loop.history = [message for message in history
                        if "[全局上下文]" not in (message.get("content") or "")]
        removed_legacy_l0 = len(loop.history) != len(history)
        loop.thread.id = data.get("thread_id")
        loop.thread.type = data.get("_thread_type")
        # 验证磁盘上线程是否仍活跃（可能已被其他进程或手动操作结束）
        if loop.thread.id and thread_store:
            thread = thread_store.load(loop.thread.id)
            if not thread or thread.get("status") != "active":
                loop.thread.id = None
                loop.thread.type = None
        loop.thread.pending_start = data.get("_pending_thread_start")
        loop.thread.current_turn_user_idx = data.get("_current_turn_user_idx", -1)
        loop.modified_values = data.get("modified_values", {})
        loop.session_id = data.get("_session_id", loop.session_id)
        loop._compressed_history_count = data.get("_compressed_history_count", 0)
        loop._session_summary = str(data.get("_session_summary") or "")
        # 旧运行态没有明确边界时，可由已归档消息数无损推导；旧的账号级
        # global_context 不迁移，避免把其他会话或设备的内容混入当前会话。
        loop._compressed_until_sequence = data.get(
            "_compressed_until_sequence", loop._compressed_history_count - 1,
        )
        # 冷存储路径：保持和 session_id 一致（按用户隔离）
        _uid = thread_store._uid() if thread_store else ""
        _cold_dir = Path(store.path).parent / "_history" / (_uid or "_anonymous")
        _cold_dir.mkdir(parents=True, exist_ok=True)
        loop._cold_store_path = _cold_dir / f"{loop.session_id}.jsonl"
        loop.child.is_child = data.get("_is_child_agent", False)
        loop.child.is_legacy = data.get("_is_legacy", False)
        loop.child.exp_id = data.get("_child_exp_id")
        loop.child.initial_history_len = data.get("_child_initial_history_len", 0)
        loop.child.agent_role = data.get("_child_agent_role")

        # 旧版本可能已将被 system 消息打断的 tool-call 批次持久化。
        # 恢复时立即规范化并写回，避免下一次请求继续触发 400。
        repaired_history = loop._repair_tool_history()

        if repaired_history or removed_legacy_l0:
            loop._persist_repaired_state()

        return loop

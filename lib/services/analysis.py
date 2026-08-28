"""跨实验分析服务：分析 Worker 直接读取选中实验的证据快照。"""

from __future__ import annotations

import json
import time
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from lib.core.prompts import ANALYSIS_SYSTEM_PROMPT


DEFAULT_ANALYSIS_REQUEST_TIMEOUT_SECONDS = 8 * 60


class AnalysisService:
    def __init__(
        self,
        exp_repo: Any,
        analysis_repo: Any,
        analyze_llm: Any,
        timeout_seconds: int | Callable[[], int] = DEFAULT_ANALYSIS_REQUEST_TIMEOUT_SECONDS,
        update_log_repo: Any = None,
        attachment_store: Any = None,
    ):
        self.exp_repo = exp_repo
        self.analysis_repo = analysis_repo
        self.analyze_llm = analyze_llm
        self.timeout_seconds = timeout_seconds
        self.update_log_repo = update_log_repo
        self.attachment_store = attachment_store

    def request_timeout_seconds(self) -> int:
        value = self.timeout_seconds() if callable(self.timeout_seconds) else self.timeout_seconds
        try:
            return max(60, min(int(value), 30 * 60))
        except (TypeError, ValueError):
            return DEFAULT_ANALYSIS_REQUEST_TIMEOUT_SECONDS

    def run_analysis(self, query: str, refs: list[str], analysis_id: str | None = None) -> dict[str, Any]:
        """建立不可变输入快照 → Worker 写报告 → 保存来源。"""
        deadline = time.monotonic() + self.request_timeout_seconds()
        packet, source_snapshot = self._build_source_packet(refs, deadline)
        remaining = int(deadline - time.monotonic())
        if remaining <= 0:
            raise TimeoutError("分析任务在读取证据时达到时限")
        analysis = self._analyze_experiments(packet, query, request_timeout=remaining)
        anal_id = self.analysis_repo.save({
            "id": analysis_id or "",
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "question": query,
            "selected_ids": refs,
            "analysis": analysis,
            "source_snapshot": source_snapshot,
        })
        for exp_id in refs:
            exp = self.exp_repo.load(exp_id)
            if exp:
                analyzed = exp.get("analyzed_in", [])
                if anal_id not in analyzed:
                    analyzed.append(anal_id)
                    exp["analyzed_in"] = analyzed
                    self.exp_repo.save(exp)
        return {
            "anal_id": anal_id,
            # 报告正文通常以“事实呈现”开头，不能拿它充当报告主题。
            # 主题就是用户确认后交给 Worker 的分析问题。
            "title": query,
            "topic": query,
            "refs": refs,
            "analysis": analysis,
            "source_snapshot": source_snapshot,
        }

    def _build_source_packet(self, refs: list[str], deadline: float | None = None) -> tuple[dict[str, Any], dict[str, Any]]:
        records: list[dict[str, Any]] = []
        updates: dict[str, list[dict[str, Any]]] = {}
        attachment_contents: list[dict[str, Any]] = []
        seen_attachments: set[str] = set()

        for exp_id in refs:
            self._ensure_before_deadline(deadline)
            experiment = self.exp_repo.load(exp_id)
            if not experiment:
                raise ValueError(f"实验不存在或已不可读取：{exp_id}")
            record = deepcopy(experiment)
            records.append(record)
            if self.update_log_repo is not None:
                updates[exp_id] = deepcopy(self.update_log_repo.list_all(exp_id))
            for link in record.get("attachments") or []:
                if not isinstance(link, dict):
                    continue
                sha256 = str(link.get("sha256") or "").strip()
                if not sha256 or sha256 in seen_attachments:
                    continue
                seen_attachments.add(sha256)
                self._ensure_before_deadline(deadline)
                attachment = self._read_attachment_for_worker(sha256, link)
                attachment_contents.append(attachment)

        snapshot = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "records": deepcopy(records),
            "update_logs": deepcopy(updates),
            # 快照保留 Worker 实际获得的附件正文，报告日后才能追溯其证据，
            # 而不是只知道当时曾有一个同名文件。
            "attachments": deepcopy(attachment_contents),
        }
        return {
            "records": records,
            "update_logs": updates,
            "attachments": attachment_contents,
        }, snapshot

    @staticmethod
    def _ensure_before_deadline(deadline: float | None) -> None:
        if deadline is not None and time.monotonic() >= deadline:
            raise TimeoutError("分析任务在读取证据时达到时限")

    def _read_attachment_for_worker(self, sha256: str, link: dict[str, Any]) -> dict[str, Any]:
        """直接读取所选实验附件，不经主 Agent 的对话上下文。"""
        base = {
            "sha256": sha256,
            "title": link.get("title", ""),
            "description": link.get("description", ""),
        }
        if self.attachment_store is None:
            return {**base, "available": False, "message": "附件存储未配置"}
        meta = self.attachment_store.meta(sha256)
        content = self.attachment_store.get(sha256)
        if not meta or content is None:
            return {**base, "available": False, "message": "附件不存在或尚未同步到本机"}
        name = str(meta.get("name") or "")
        mime = str(meta.get("mime") or "")
        suffix = Path(name).suffix.lower()
        result = {
            **base,
            "available": True,
            "name": name,
            "mime": mime,
            "size": len(content),
            "extraction": "metadata_only",
        }
        try:
            if mime.startswith("text/") or suffix in (".csv", ".txt", ".tsv"):
                result.update({
                    "extraction": "text",
                    "content": content.decode("utf-8-sig", errors="replace"),
                })
            elif suffix == ".xlsx" or mime == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ):
                from openpyxl import load_workbook
                book = load_workbook(BytesIO(content), read_only=True, data_only=True)
                sheets = []
                for sheet in book.worksheets:
                    rows = [
                        "\t".join("" if value is None else str(value) for value in row)
                        for row in sheet.iter_rows(values_only=True)
                    ]
                    sheets.append({
                        "name": sheet.title,
                        "rows": sheet.max_row or 0,
                        "columns": sheet.max_column or 0,
                        "content": "\n".join(rows),
                    })
                book.close()
                result.update({"extraction": "xlsx", "sheets": sheets})
            elif suffix == ".pdf" or mime == "application/pdf":
                from pypdf import PdfReader
                reader = PdfReader(BytesIO(content))
                result.update({
                    "extraction": "pdf",
                    "content": "\n\n".join(page.extract_text() or "" for page in reader.pages),
                })
            elif mime.startswith("image/") or suffix in (
                ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"
            ):
                # OCR 已废弃：图片内容由主 Agent 的视觉模型阅读，分析 Worker 不再执行本地 OCR。
                result["message"] = "图片正文依赖主 Agent 视觉识别；分析 Worker 不执行本地 OCR。"
            else:
                result["message"] = "此文件类型暂不支持提取正文。"
        except Exception as exc:
            result.update({
                "available": False,
                "message": f"无法提取附件正文：{str(exc)[:160]}",
            })
        return result

    def _analyze_experiments(
        self, packet: dict[str, Any], question: str, request_timeout: int | None = None
    ) -> str:
        evidence = json.dumps(packet, ensure_ascii=False, indent=2, default=str)
        user_prompt = f"""RESEARCHER'S QUESTION:
{question}

EVIDENCE SNAPSHOT:
{evidence}

Use the complete experiment records, their update logs, and the extracted
contents of their linked attachments as evidence. Do not claim to have read an
attachment marked unavailable or metadata_only. Cite experiment IDs and, when
used, attachment names and worksheet/page context in the report.
Treat the researcher's question as the report scope. Do not add a data-quality
audit, a copy-trace check, or another analysis dimension unless the question
explicitly asks for it.

Structure your response in exactly three sections as specified in the system
prompt: 事实呈现, 发现提示, 值得思考的问题."""
        llm = self.analyze_llm() if callable(self.analyze_llm) else self.analyze_llm
        if llm is None:
            raise RuntimeError("未配置分析模型 API Key")
        return str(llm.analyze(
            ANALYSIS_SYSTEM_PROMPT,
            user_prompt,
            request_timeout=request_timeout or self.request_timeout_seconds(),
            max_attempts=1,
        ))
